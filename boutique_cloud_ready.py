import streamlit as st
from pymongo import MongoClient
import pandas as pd
from datetime import date, datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import hashlib
import json
import math
import os
import re
import time
import hmac
import uuid
from html import escape as html_escape
import bcrypt
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill

load_dotenv("credentials/.env")

# =====================================================
# PASSWORD HASH UTILITY
# Run once in a Python shell to generate your hash:
#
#   import bcrypt
#   h = bcrypt.hashpw(b"your_password_here", bcrypt.gensalt())
#   print(h.decode())
#
# Then set PASSWORD_HASH=<output> in credentials/.env
# or in your Streamlit secrets.toml as:
#   PASSWORD_HASH = "<output>"
# =====================================================

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="Vinay Boutique",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =====================================================
# CSS — DARK NAVY BLUE THEME
# =====================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;1,400&family=DM+Serif+Display:ital@0;1&display=swap');

:root {
    /* Light mode palette */
    --bg:           #F4F7FC;
    --bg-2:         #EBF0FA;
    --surface:      #FFFFFF;
    --surface-2:    #F8FAFF;
    --blue:         #2563EB;
    --blue-soft:    #3B82F6;
    --blue-pale:    #BFDBFE;
    --blue-glow:    rgba(37,99,235,0.12);
    --text:         #0F172A;
    --text-2:       #334155;
    --muted:        #64748B;
    --dim:          #94A3B8;
    --emerald:      #059669;
    --rose:         #DC2626;
    --amber:        #D97706;
    --r:            10px;
    --r-lg:         14px;
    --r-xl:         22px;
    --border:       rgba(37,99,235,0.12);
    --border-hover: rgba(37,99,235,0.35);
    --shadow-sm:    0 1px 3px rgba(15,23,42,0.06), 0 1px 2px rgba(15,23,42,0.04);
    --shadow:       0 4px 16px rgba(15,23,42,0.08), 0 2px 6px rgba(15,23,42,0.05);
    --shadow-lg:    0 12px 40px rgba(37,99,235,0.14), 0 4px 12px rgba(15,23,42,0.06);
    /* Legacy aliases kept so existing rules don't break */
    --navy-1:       #F4F7FC;
    --navy-2:       #EBF0FA;
    --navy-3:       #FFFFFF;
    --navy-4:       #F0F5FF;
    --navy-5:       #E2EAFF;
    --cream:        #0F172A;
    --cream-dim:    #334155;
}

/* ━━━ SIDEBAR THEME TOGGLE ━━━ */
[data-testid="stSidebar"] > div:first-child > div:first-child .stButton > button {
    background: transparent !important;
    border: 1px solid var(--border) !important;
    color: var(--muted) !important;
    font-size: 0.7rem !important;
    padding: 0.3rem 0.8rem !important;
    letter-spacing: 0.06em !important;
    width: auto !important;
    float: right;
    margin-bottom: 0.5rem;
}
[data-testid="stSidebar"] > div:first-child > div:first-child .stButton > button:hover {
    border-color: var(--border-hover) !important;
    color: var(--blue) !important;
    transform: none !important;
    background: var(--blue-glow) !important;
}

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif !important;
    background: var(--bg) !important;
    color: var(--text) !important;
}

/* ━━━ ALL INPUTS — DARK TEXT FOR LIGHT MODE ━━━ */
input:not([type="radio"]):not([type="checkbox"]),
textarea,
[data-baseweb="input"] input,
[data-baseweb="base-input"] input,
[data-baseweb="textarea"] textarea,
[data-baseweb="date-picker"] input,
[data-baseweb="select"] input {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
    caret-color: var(--blue) !important;
    background-color: var(--surface) !important;
}
[data-baseweb="input"] *,
[data-baseweb="base-input"] *,
[data-baseweb="textarea"] * {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Selectbox displayed value */
[data-baseweb="select"] > div > div,
[data-baseweb="select"] > div > div > div,
[data-baseweb="select"] span,
[class*="ValueContainer"] > div,
[class*="singleValue"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Number stepper */
.stNumberInput div[data-baseweb="input"] input,
.stNumberInput input[type="number"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
/* Date displayed value */
.stDateInput div[data-baseweb="input"] input,
.stDateInput input[type="text"] {
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stApp {
    background: var(--bg) !important;
    background-image:
        radial-gradient(ellipse 900px 600px at 0% 0%, rgba(37,99,235,0.04) 0%, transparent 70%),
        radial-gradient(ellipse 700px 500px at 100% 100%, rgba(37,99,235,0.03) 0%, transparent 70%);
    background-attachment: fixed;
}
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: var(--bg-2); border-radius: 99px; }
::-webkit-scrollbar-thumb { background: rgba(37,99,235,0.3); border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: rgba(37,99,235,0.5); }

h1, h2, h3 {
    font-family: 'DM Serif Display', serif !important;
    color: var(--text) !important;
    letter-spacing: -0.01em;
    font-weight: 400 !important;
}
h4, h5, h6 {
    font-family: 'DM Sans', sans-serif !important;
    color: var(--muted) !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-size: 0.7rem !important;
    font-weight: 600 !important;
}

.page-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2.1rem;
    font-weight: 400;
    color: var(--text);
    letter-spacing: -0.02em;
    line-height: 1.15;
    margin-bottom: 0.2rem;
}
.page-sub {
    font-size: 0.7rem;
    text-transform: uppercase;
    letter-spacing: 0.18em;
    color: var(--dim);
    margin-bottom: 2.4rem;
    font-weight: 500;
}
.rule { height:1px; background:linear-gradient(90deg, var(--blue) 0%, rgba(37,99,235,0.15) 60%, transparent 100%); margin:2rem 0; border:none; }
.rule-sm { height:1px; background:linear-gradient(90deg, rgba(37,99,235,0.25), transparent); margin:1.2rem 0; border:none; }

/* ━━━ SIDEBAR ━━━ */
[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
    box-shadow: 2px 0 12px rgba(15,23,42,0.05) !important;
}
[data-testid="stSidebar"] * { color: var(--text) !important; }
.sb-brand { padding: 2rem 1.5rem 0.5rem; text-align: center; }
.sb-logo {
    font-family: 'DM Serif Display', serif;
    font-size: 1.75rem;
    font-weight: 400;
    color: var(--blue);
    letter-spacing: -0.01em;
    line-height: 1;
}
.sb-mark {
    font-size: 0.6rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--muted);
    margin-top: 0.3rem;
    font-weight: 600;
}
[data-testid="stSidebar"] .stRadio > div {
    gap: 2px !important;
    flex-direction: column !important;
}
[data-testid="stSidebar"] .stRadio > div > label {
    background: transparent !important;
    border: none !important;
    border-radius: var(--r) !important;
    color: var(--muted) !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.01em !important;
    padding: 0.55rem 1rem 0.55rem 0.75rem !important;
    transition: all 0.18s ease !important;
    cursor: pointer;
    display: flex;
    align-items: center;
}
[data-testid="stSidebar"] .stRadio > div > label:hover {
    background: var(--blue-glow) !important;
    color: var(--blue) !important;
}
[data-testid="stSidebar"] .stRadio > div > label > div:first-child {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 16px !important;
    height: 16px !important;
    min-width: 16px !important;
    border: 2px solid var(--dim) !important;
    border-radius: 50% !important;
    margin-right: 8px !important;
    background: transparent !important;
    transition: all 0.18s !important;
}
.sb-user {
    font-size: 0.75rem;
    color: var(--muted);
    text-align: center;
    padding: 0.6rem 0 1.5rem;
    letter-spacing: 0.04em;
    font-weight: 500;
}
.sb-sep { height: 1px; background: var(--border); margin: 0.8rem 1rem; }

/* ━━━ PUBLIC BANNER ━━━ */
.pub-banner {
    background: linear-gradient(135deg, #FFFFFF 0%, #F0F6FF 100%);
    border: 1px solid var(--border);
    border-radius: var(--r-xl);
    padding: 2.2rem 2.8rem 1.8rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.pub-banner::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft), transparent);
    border-radius: var(--r-xl) var(--r-xl) 0 0;
}
.pub-banner-title {
    font-family: 'DM Serif Display', serif;
    font-size: 2rem;
    font-weight: 400;
    color: var(--text);
    letter-spacing: -0.02em;
    line-height: 1.1;
    margin-bottom: 0.3rem;
}
.pub-banner-sub {
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--muted);
    font-weight: 600;
}

/* ━━━ ADMIN LOGIN PANEL (bottom) ━━━ */
.admin-strip {
    margin-top: 3rem;
    border-top: 1px solid var(--border);
    padding-top: 1.5rem;
}
.admin-strip-label {
    font-size: 0.62rem;
    text-transform: uppercase;
    letter-spacing: 0.22em;
    color: var(--dim);
    text-align: center;
    margin-bottom: 0.8rem;
    font-weight: 600;
}

/* ━━━ METRICS ━━━ */
[data-testid="stMetric"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r-lg) !important;
    padding: 1.25rem 1.4rem !important;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow-sm) !important;
    transition: box-shadow 0.2s ease, transform 0.2s ease !important;
}
[data-testid="stMetric"]::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft), transparent);
    opacity: 0.6;
}
[data-testid="stMetric"]:hover {
    box-shadow: var(--shadow) !important;
    transform: translateY(-2px);
}
[data-testid="stMetricLabel"] > div {
    color: var(--muted) !important;
    font-size: 0.68rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.12em !important;
    font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stMetricValue"] {
    color: var(--text) !important;
    font-family: 'DM Serif Display', serif !important;
    font-size: 1.7rem !important;
    font-weight: 400 !important;
    letter-spacing: -0.02em !important;
    line-height: 1.2 !important;
}

/* ━━━ BUTTONS ━━━ */
.stButton > button {
    background: var(--surface) !important;
    color: var(--blue) !important;
    border: 1.5px solid rgba(37,99,235,0.28) !important;
    border-radius: var(--r) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.6rem 1.4rem !important;
    transition: all 0.18s ease !important;
    box-shadow: var(--shadow-sm) !important;
    width: 100% !important;
}
.stButton > button:hover {
    background: var(--blue) !important;
    border-color: var(--blue) !important;
    color: #FFFFFF !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 16px rgba(37,99,235,0.25) !important;
}
.stButton > button:active { transform: scale(0.98) !important; }

.stDownloadButton > button {
    background: var(--surface-2) !important;
    color: var(--muted) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.04em !important;
    transition: all 0.18s ease !important;
    width: 100% !important;
}
.stDownloadButton > button:hover {
    border-color: var(--border-hover) !important;
    color: var(--blue) !important;
    background: var(--blue-glow) !important;
}

/* ━━━ FORM SUBMIT ━━━ */
.stForm button[type="submit"] {
    background: var(--blue) !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: var(--r) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.84rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.85rem 2.5rem !important;
    width: 100% !important;
    transition: all 0.2s ease !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.2) !important;
}
.stForm button[type="submit"]:hover {
    background: #1D4ED8 !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 24px rgba(37,99,235,0.35) !important;
}

/* ━━━ INPUTS ━━━ */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea,
.stDateInput > div > div > input,
.stDateInput input,
input[type="text"], input[type="number"], input[type="date"], textarea {
    background: var(--surface) !important;
    border: 1.5px solid rgba(37,99,235,0.18) !important;
    border-radius: var(--r) !important;
    color: var(--text) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.88rem !important;
    font-weight: 400 !important;
    padding: 0.6rem 0.9rem !important;
    transition: border-color 0.18s ease, box-shadow 0.18s ease !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus,
.stDateInput > div > div > input:focus {
    border-color: var(--blue) !important;
    box-shadow: 0 0 0 3px rgba(37,99,235,0.1) !important;
    color: var(--text) !important;
    -webkit-text-fill-color: var(--text) !important;
}
.stTextInput > div > div > input::placeholder,
.stTextArea > div > div > textarea::placeholder,
.stDateInput > div > div > input::placeholder { color: var(--dim) !important; -webkit-text-fill-color: var(--dim) !important; }

/* Date picker calendar popup */
[data-baseweb="calendar"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    box-shadow: var(--shadow-lg) !important;
    border-radius: var(--r-lg) !important;
}
[data-baseweb="calendar"] * { color: var(--text) !important; background: transparent !important; }
[data-baseweb="calendar"] [aria-selected="true"] { background: var(--blue) !important; color: #FFFFFF !important; }
[data-baseweb="calendar"] button:hover { background: var(--bg-2) !important; }

/* Date input wrapper */
.stDateInput > div {
    background: var(--surface) !important;
    border-radius: var(--r) !important;
}
.stDateInput > div > div { background: var(--surface) !important; }
.stDateInput svg { fill: var(--muted) !important; }

/* Selectbox */
.stSelectbox > div > div,
.stSelectbox [data-baseweb="select"] > div {
    background: var(--surface) !important;
    border: 1.5px solid rgba(37,99,235,0.18) !important;
    border-radius: var(--r) !important;
    color: var(--text) !important;
    transition: border-color 0.18s ease !important;
}
.stSelectbox > div > div:hover,
.stSelectbox [data-baseweb="select"] > div:hover { border-color: var(--border-hover) !important; }
.stSelectbox [data-baseweb="select"] span,
.stSelectbox [data-baseweb="select"] div { color: var(--text) !important; }

/* Selectbox dropdown menu */
[data-baseweb="popover"] [data-baseweb="menu"],
[data-baseweb="popover"] ul {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    box-shadow: var(--shadow-lg) !important;
    border-radius: var(--r-lg) !important;
}
[data-baseweb="popover"] li,
[data-baseweb="popover"] [role="option"] {
    background: var(--surface) !important;
    color: var(--text) !important;
}
[data-baseweb="popover"] li:hover,
[data-baseweb="popover"] [role="option"]:hover { background: var(--bg-2) !important; }

/* Number input */
.stNumberInput > div > div { background: var(--surface) !important; border-radius: var(--r) !important; }
.stNumberInput input { color: var(--text) !important; -webkit-text-fill-color: var(--text) !important; }

/* ━━━ LABELS ━━━ */
.stTextInput label, .stNumberInput label, .stSelectbox label,
.stTextArea label, .stDateInput label, .stRadio label, .stCheckbox label,
.stTextInput label p, .stNumberInput label p, .stSelectbox label p,
.stTextArea label p, .stDateInput label p, .stRadio label p,
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span {
    color: var(--text-2) !important;
    -webkit-text-fill-color: var(--text-2) !important;
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-family: 'DM Sans', sans-serif !important;
}

/* ━━━ DATAFRAME ━━━ */
.stDataFrame {
    border-radius: var(--r-lg) !important;
    border: 1px solid var(--border) !important;
    overflow: hidden !important;
    box-shadow: var(--shadow-sm) !important;
}
[data-testid="stDataFrame"] th {
    background: var(--bg-2) !important;
    color: var(--muted) !important;
    font-size: 0.68rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.1em !important;
    font-weight: 700 !important;
    border-bottom: 1px solid var(--border) !important;
    padding: 0.75rem 1rem !important;
}
[data-testid="stDataFrame"] td {
    background: var(--surface) !important;
    color: var(--text-2) !important;
    font-size: 0.85rem !important;
    font-weight: 400 !important;
    border-bottom: 1px solid rgba(37,99,235,0.06) !important;
    padding: 0.7rem 1rem !important;
}
[data-testid="stDataFrame"] tr:hover td { background: var(--bg-2) !important; }

/* ━━━ TABS ━━━ */
.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 2px solid var(--border) !important;
    border-radius: 0 !important;
    padding: 0 !important;
    gap: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 0 !important;
    color: var(--dim) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.04em !important;
    padding: 0.75rem 1.25rem !important;
    border-bottom: 2px solid transparent !important;
    transition: all 0.18s !important;
    margin-bottom: -2px !important;
}
.stTabs [aria-selected="true"] {
    color: var(--blue) !important;
    border-bottom-color: var(--blue) !important;
    font-weight: 600 !important;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--text) !important; }

/* ━━━ ALERTS ━━━ */
.stSuccess {
    background: rgba(5,150,105,0.07) !important;
    border: 1px solid rgba(5,150,105,0.25) !important;
    border-radius: var(--r) !important;
}
.stSuccess * { color: #065f46 !important; }
.stInfo {
    background: rgba(37,99,235,0.06) !important;
    border: 1px solid rgba(37,99,235,0.2) !important;
    border-radius: var(--r) !important;
}
.stInfo * { color: #1e3a8a !important; }
.stWarning {
    background: rgba(217,119,6,0.07) !important;
    border: 1px solid rgba(217,119,6,0.25) !important;
    border-radius: var(--r) !important;
}
.stWarning * { color: #92400e !important; }
.stError {
    background: rgba(220,38,38,0.06) !important;
    border: 1px solid rgba(220,38,38,0.22) !important;
    border-radius: var(--r) !important;
}
.stError * { color: #991b1b !important; }

/* ━━━ RADIO (inline) ━━━ */
.stRadio > div { gap: 0.5rem !important; flex-direction: row !important; }
.stRadio > div > label {
    background: var(--surface) !important;
    border: 1.5px solid var(--border) !important;
    border-radius: var(--r) !important;
    padding: 0.5rem 1.1rem !important;
    color: var(--muted) !important;
    font-size: 0.8rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.02em !important;
    transition: all 0.18s !important;
    cursor: pointer;
    box-shadow: var(--shadow-sm) !important;
}
.stRadio > div > label:hover {
    border-color: var(--blue) !important;
    color: var(--blue) !important;
    background: var(--blue-glow) !important;
}

/* ━━━ SECTION HEADERS ━━━ */
.sec-head {
    font-family: 'DM Serif Display', serif;
    font-size: 1.05rem;
    font-weight: 400;
    font-style: italic;
    color: var(--text);
    margin: 1.8rem 0 1rem;
    padding-bottom: 0.6rem;
    border-bottom: 1px solid var(--border);
    letter-spacing: -0.01em;
}

/* ━━━ BADGES ━━━ */
.badge {
    display: inline-block;
    font-size: 0.68rem;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    padding: 0.22rem 0.65rem;
    border-radius: 6px;
}
.badge-gold  { background: rgba(37,99,235,0.1); color: var(--blue); }
.badge-green { background: rgba(5,150,105,0.1); color: #065f46; }
.badge-red   { background: rgba(220,38,38,0.09); color: #991b1b; }
.badge-muted { background: var(--bg-2); color: var(--muted); }

/* ━━━ EMPTY STATE ━━━ */
.empty { text-align: center; padding: 4rem 2rem; color: var(--dim); }
.empty-glyph { font-size: 2rem; margin-bottom: 1rem; color: var(--border); }

/* Number input spinners */
button[data-testid="stNumberInputStepDown"],
button[data-testid="stNumberInputStepUp"] {
    background: var(--bg-2) !important;
    border-color: var(--border) !important;
    color: var(--muted) !important;
}

/* Expanders */
.streamlit-expanderHeader {
    background: var(--surface-2) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    color: var(--text-2) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.84rem !important;
    font-weight: 500 !important;
}
.streamlit-expanderContent {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-top: none !important;
    border-radius: 0 0 var(--r) var(--r) !important;
    padding: 1rem !important;
}

/* ━━━ BACKUP & RESTORE CARDS ━━━ */
.bk-card {
    background: var(--surface);
    border: 1.5px solid var(--border);
    border-radius: var(--r-xl);
    padding: 1.8rem 2rem;
    box-shadow: var(--shadow);
    transition: box-shadow 0.2s ease, transform 0.2s ease;
    animation: fadeSlideUp 0.4s ease forwards;
    opacity: 0;
}
.bk-card:hover { box-shadow: var(--shadow-lg); transform: translateY(-2px); }
@keyframes fadeSlideUp {
    from { opacity: 0; transform: translateY(16px); }
    to   { opacity: 1; transform: translateY(0);   }
}
.bk-card-icon {
    width: 42px; height: 42px;
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.25rem;
    margin-bottom: 1rem;
}
.bk-icon-blue  { background: rgba(37,99,235,0.1); }
.bk-icon-green { background: rgba(5,150,105,0.1); }
.bk-icon-amber { background: rgba(217,119,6,0.1); }
.bk-card-title {
    font-family: 'DM Sans', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: var(--text);
    margin-bottom: 0.2rem;
    letter-spacing: -0.01em;
}
.bk-card-desc {
    font-size: 0.78rem;
    color: var(--muted);
    margin-bottom: 1.2rem;
    line-height: 1.5;
}
.bk-status-badge {
    display: inline-flex; align-items: center; gap: 0.4rem;
    font-size: 0.7rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.1em;
    padding: 0.28rem 0.7rem;
    border-radius: 999px;
}
.bk-status-ok    { background: rgba(5,150,105,0.1); color: #065f46; }
.bk-status-warn  { background: rgba(217,119,6,0.1);  color: #92400e; }
.bk-status-info  { background: rgba(37,99,235,0.1);  color: #1e3a8a; }
.bk-header {
    background: linear-gradient(135deg, #FFFFFF 0%, #EBF3FF 100%);
    border: 1.5px solid var(--border);
    border-radius: var(--r-xl);
    padding: 2rem 2.4rem;
    margin-bottom: 1.8rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow);
}
.bk-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft));
    border-radius: var(--r-xl) var(--r-xl) 0 0;
}
.bk-header-icon {
    font-size: 2rem; margin-bottom: 0.6rem; display: block;
}
.bk-header-title {
    font-family: 'DM Serif Display', serif;
    font-size: 1.75rem; font-weight: 400; color: var(--text);
    letter-spacing: -0.02em; line-height: 1.1; margin-bottom: 0.3rem;
}
.bk-header-sub {
    font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.16em;
    color: var(--muted); font-weight: 600;
}
.bk-progress-bar {
    height: 6px; border-radius: 999px;
    background: linear-gradient(90deg, var(--blue), var(--blue-soft));
    animation: progressPulse 1.5s ease-in-out infinite alternate;
}
@keyframes progressPulse {
    from { opacity: 0.6; width: 30%; }
    to   { opacity: 1;   width: 85%; }
}
.bk-ts {
    font-size: 0.72rem; color: var(--muted);
    font-family: 'DM Sans', sans-serif; margin-top: 0.5rem;
}

/* File uploader */
[data-testid="stFileUploader"] > div {
    background: var(--surface-2) !important;
    border: 2px dashed rgba(37,99,235,0.25) !important;
    border-radius: var(--r-lg) !important;
    transition: border-color 0.18s !important;
}
[data-testid="stFileUploader"] > div:hover {
    border-color: var(--blue) !important;
    background: var(--blue-glow) !important;
}
[data-testid="stFileUploader"] * { color: var(--muted) !important; }

.stCaption, .stCaption * { color: var(--dim) !important; }

p, span { color: var(--text-2); }

.skeleton {
    min-height: 72px;
    border-radius: var(--r-lg);
    background: linear-gradient(90deg, var(--surface-2) 0%, rgba(37,99,235,0.08) 45%, var(--surface-2) 90%);
    background-size: 220% 100%;
    animation: skeletonSweep 1.2s ease-in-out infinite;
    border: 1px solid var(--border);
}
@keyframes skeletonSweep {
    from { background-position: 0% 0; }
    to { background-position: 220% 0; }
}

@media (max-width: 760px) {
    .block-container { padding-left: 1rem !important; padding-right: 1rem !important; }
    [data-testid="stHorizontalBlock"] {
        flex-wrap: wrap !important;
        gap: 0.75rem !important;
    }
    [data-testid="column"] {
        flex: 1 1 100% !important;
        min-width: 100% !important;
        width: 100% !important;
    }
    [data-testid="stMetric"] {
        padding: 0.9rem 1rem !important;
    }
    [data-testid="stMetricValue"] {
        font-size: 1.35rem !important;
        overflow-wrap: anywhere !important;
    }
    .pub-banner, .bk-header, .bk-card {
        padding: 1.2rem 1rem !important;
        border-radius: var(--r-lg) !important;
    }
    .page-title { font-size: 1.65rem !important; }
    .stTabs [data-baseweb="tab-list"] {
        overflow-x: auto !important;
        flex-wrap: nowrap !important;
    }
}
</style>
""", unsafe_allow_html=True)

# =====================================================
# PLOTLY TEMPLATE
# =====================================================


# ── THEME TOGGLE — pure session_state, no JS needed ─────────────────────────
if "theme" not in st.session_state:
    st.session_state.theme = "system"

_DARK_CSS = """
<style>
/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   DARK MODE OVERRIDE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
:root {
    --bg:           #070C18 !important;
    --bg-2:         #0B1221 !important;
    --surface:      #0F1A2E !important;
    --surface-2:    #152238 !important;
    --blue:         #4D8AE8 !important;
    --blue-soft:    #6BA3F0 !important;
    --blue-pale:    #A8C4F0 !important;
    --blue-glow:    rgba(77,138,232,0.15) !important;
    --text:         #E8EEF8 !important;
    --text-2:       #C8D4E8 !important;
    --muted:        #6A84A8 !important;
    --dim:          #3D5478 !important;
    --border:       rgba(77,138,232,0.18) !important;
    --border-hover: rgba(77,138,232,0.42) !important;
    --shadow-sm:    0 1px 6px rgba(0,0,0,0.4) !important;
    --shadow:       0 4px 24px rgba(0,0,0,0.5) !important;
    --shadow-lg:    0 8px 48px rgba(0,0,0,0.65) !important;
}
.stApp, .stApp > div, .main, [data-testid="stAppViewContainer"], [data-testid="block-container"] {
    background: #070C18 !important;
}
[data-testid="stSidebar"], [data-testid="stSidebar"] > div { background: #0B1221 !important; }
[data-testid="stSidebar"] * { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
p, span, div, label { color: #C8D4E8 !important; }
h1, h2, h3 { color: #E8EEF8 !important; }
[data-testid="stMetric"] { background: #0F1A2E !important; }
[data-testid="stMetricValue"] { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
input, textarea, [data-baseweb="input"] input, [data-baseweb="base-input"] input {
    background: #0F1A2E !important; color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important;
}
.stSelectbox > div > div, .stSelectbox [data-baseweb="select"] > div, [data-baseweb="select"] > div { background: #0F1A2E !important; }
[data-baseweb="select"] span, [data-baseweb="select"] div, [class*="singleValue"] { color: #E8EEF8 !important; -webkit-text-fill-color: #E8EEF8 !important; }
[data-baseweb="popover"] [data-baseweb="menu"], [data-baseweb="popover"] ul { background: #0F1A2E !important; }
[data-baseweb="popover"] li, [data-baseweb="popover"] [role="option"] { background: #0F1A2E !important; color: #E8EEF8 !important; }
[data-baseweb="popover"] li:hover, [data-baseweb="popover"] [role="option"]:hover { background: #1C2D47 !important; }
[data-baseweb="calendar"] { background: #0F1A2E !important; }
[data-baseweb="calendar"] * { color: #E8EEF8 !important; }
.stDateInput > div, .stDateInput > div > div { background: #0F1A2E !important; }
[data-testid="stDataFrame"] th { background: #0F1A2E !important; color: #6A84A8 !important; }
[data-testid="stDataFrame"] td { background: #0B1221 !important; color: #C8D4E8 !important; }
[data-testid="stDataFrame"] tr:hover td { background: #0F1A2E !important; }
.stButton > button { background: transparent !important; color: #4D8AE8 !important; border-color: rgba(77,138,232,0.42) !important; }
.stButton > button:hover { background: rgba(77,138,232,0.14) !important; color: #A8C4F0 !important; }
.stDownloadButton > button { background: transparent !important; color: #6A84A8 !important; }
.stTabs [data-baseweb="tab"] { color: #3D5478 !important; }
.stTabs [aria-selected="true"] { color: #E8EEF8 !important; }
.streamlit-expanderHeader { background: #0F1A2E !important; color: #6A84A8 !important; border-color: rgba(77,138,232,0.18) !important; }
.streamlit-expanderContent { background: #0B1221 !important; border-color: rgba(77,138,232,0.18) !important; }
.stRadio > div > label { background: #0F1A2E !important; color: #6A84A8 !important; border-color: rgba(77,138,232,0.18) !important; }
.pub-banner { background: linear-gradient(135deg, #0B1221 0%, #0F1A2E 100%) !important; }
.pub-banner-title { color: #E8EEF8 !important; }
.bk-card { background: #0F1A2E !important; }
.bk-header { background: linear-gradient(135deg, #0B1221 0%, #0F1A2E 100%) !important; }
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p, [data-testid="stWidgetLabel"] span { color: #C8D4E8 !important; -webkit-text-fill-color: #C8D4E8 !important; }
.stSuccess { background: rgba(61,154,108,0.12) !important; } .stSuccess * { color: #7ADFA0 !important; }
.stInfo { background: rgba(77,138,232,0.1) !important; } .stInfo * { color: #A8C4F0 !important; }
.stWarning { background: rgba(200,160,50,0.1) !important; } .stWarning * { color: #E8C840 !important; }
.stError { background: rgba(192,80,96,0.1) !important; } .stError * { color: #E08090 !important; }
[data-testid="stFileUploader"] > div { background: #0F1A2E !important; border-color: rgba(77,138,232,0.3) !important; }
</style>
"""

_SYSTEM_DARK_CSS = _DARK_CSS.replace("<style>", "<style>@media (prefers-color-scheme: dark) {", 1).replace("</style>", "}</style>", 1)

def inject_theme():
    if st.session_state.theme == "dark":
        st.markdown(_DARK_CSS, unsafe_allow_html=True)
    elif st.session_state.theme == "system":
        st.markdown(_SYSTEM_DARK_CSS, unsafe_allow_html=True)

PLOT_LAYOUT = dict(
    paper_bgcolor="rgba(255,255,255,0)",
    plot_bgcolor="rgba(255,255,255,0)",
    font=dict(family="DM Sans", color="#64748B", size=11),
    title=dict(font=dict(family="DM Serif Display", size=17, color="#0F172A"), pad=dict(b=12), x=0),
    xaxis=dict(gridcolor="rgba(37,99,235,0.08)", linecolor="rgba(37,99,235,0.15)", tickfont=dict(size=10, color="#64748B"), showgrid=True, zeroline=False),
    yaxis=dict(gridcolor="rgba(37,99,235,0.08)", linecolor="rgba(37,99,235,0.15)", tickfont=dict(size=10, color="#64748B"), showgrid=True, zeroline=False),
    legend=dict(bgcolor="rgba(255,255,255,0.92)", bordercolor="rgba(37,99,235,0.15)", borderwidth=1, font=dict(color="#64748B", size=10)),
    margin=dict(l=12, r=12, t=44, b=12),
    colorway=["#2563EB","#3B82F6","#059669","#8BACD8","#DC2626","#1D4ED8","#10B981","#0EA5E9"],
    hoverlabel=dict(bgcolor="rgba(255,255,255,0.97)", bordercolor="rgba(37,99,235,0.3)", font=dict(color="#0F172A", size=11, family="DM Sans"), align="left"),
    bargap=0.35,
)

_DARK_PLOT_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="DM Sans", color="#3D5478", size=11),
    title=dict(font=dict(family="DM Serif Display", size=17, color="#C8D4E8"), pad=dict(b=12), x=0),
    xaxis=dict(gridcolor="rgba(77,138,232,0.08)", linecolor="rgba(77,138,232,0.15)", tickfont=dict(size=10, color="#3D5478"), showgrid=True, zeroline=False),
    yaxis=dict(gridcolor="rgba(77,138,232,0.08)", linecolor="rgba(77,138,232,0.15)", tickfont=dict(size=10, color="#3D5478"), showgrid=True, zeroline=False),
    legend=dict(bgcolor="rgba(7,12,24,0.88)", bordercolor="rgba(77,138,232,0.2)", borderwidth=1, font=dict(color="#6A84A8", size=10)),
    margin=dict(l=12, r=12, t=44, b=12),
    colorway=["#4D8AE8","#6BA3F0","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8"],
    hoverlabel=dict(bgcolor="rgba(7,12,24,0.96)", bordercolor="rgba(77,138,232,0.3)", font=dict(color="#E8EEF8", size=11, family="DM Sans"), align="left"),
    bargap=0.35,
)

def get_plot_layout():
    """Return plotly layout dict adjusted for current theme."""
    if st.session_state.get("theme") == "dark":
        return _DARK_PLOT_LAYOUT
    return PLOT_LAYOUT

def styled_fig(fig, height=340):
    fig.update_layout(**get_plot_layout(), height=height)
    return fig

# =====================================================
# CONSTANTS
# =====================================================

CATEGORIES      = ["Sarees","Salwar Suits","Lehengas","Kurtis","Western Wear","Accessories","Kids Wear","Blouse","Fabric","Other"]
PAYMENT_METHODS = ["Cash","UPI","Card","Bank Transfer","Part Payment","Credit"]
TRANSACTION_TYPES = ["sale", "pending_payment", "return", "expense"]
EXPENSE_CATEGORIES = ["Rent","Electricity","Staff","Packaging","Alteration","Marketing","Transport","Vendor Payment","Other"]
SESSION_TIMEOUT_SECONDS = int(os.getenv("SESSION_TIMEOUT_SECONDS", "1800"))
STAFF_ROLES = ["cashier", "manager", "admin"]
ROLE_PAGES = {
    "cashier": ["Dashboard", "Add Sale", "Review Accounts", "Customer List", "Reminders & Alerts", "Daily Cash Summary", "Logout"],
    "manager": ["Dashboard", "Add Sale", "Review Accounts", "Update Transaction", "Customer List", "Analytics", "Reminders & Alerts", "Daily Cash Summary", "Vendor Payables", "Inventory Tracker", "Operations Hub", "Logout"],
    "admin": ["Dashboard", "Add Sale", "Review Accounts", "Update Transaction", "Customer List", "Analytics", "Reminders & Alerts", "Daily Cash Summary", "Vendor Payables", "Inventory Tracker", "Operations Hub", "Staff & Audit", "Backup & Restore", "Logout"],
}
LOYALTY_TIERS = [
    {"tier": "Bronze", "min_spend": 0, "discount_pct": 0.0, "points_rate": 1},
    {"tier": "Silver", "min_spend": 5000, "discount_pct": 2.0, "points_rate": 1},
    {"tier": "Gold", "min_spend": 20000, "discount_pct": 5.0, "points_rate": 2},
    {"tier": "Platinum", "min_spend": 50000, "discount_pct": 8.0, "points_rate": 3},
]
COUPONS = {
    "WELCOME5": {"discount_pct": 5.0, "label": "Welcome discount"},
    "FESTIVE10": {"discount_pct": 10.0, "label": "Festival campaign"},
}
ORDER_STATUSES = ["Queued", "In Progress", "Ready", "Delivered", "Cancelled"]
PAYMENT_STATUSES = ["Unpaid", "Part Paid", "Paid"]
CAMPAIGN_CHANNELS = ["WhatsApp", "SMS", "Email", "Phone"]

# =====================================================
# MONGODB
# =====================================================

@st.cache_resource
def get_mongo_client():
    try:
        try:
            uri = st.secrets.get("MONGO_URI", os.getenv("MONGO_URI"))
        except Exception:
            uri = os.getenv("MONGO_URI")
        if not uri:
            st.error("⚠️ MONGO_URI not configured.")
            st.stop()
        client = MongoClient(uri, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        return client
    except Exception as e:
        st.error(f"MongoDB connection failed: {e}")
        st.stop()

def get_db():
    return get_mongo_client()["boutique_db"]

def get_col():
    return get_db()["sales"]

def get_next_id():
    counter = get_db()["counters"].find_one_and_update(
        {"_id": "sales_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    return counter["seq"]

def get_staff_col():
    return get_db()["staff_users"]

def get_audit_col():
    return get_db()["audit_logs"]

def get_collection(name: str):
    return get_db()[name]

def active_filter(extra: dict | None = None, include_deleted: bool = False) -> dict:
    query = {} if include_deleted else {"deleted_at": {"$exists": False}}
    if extra:
        query.update(extra)
    return query

def current_user() -> str:
    return st.session_state.get("username") or "System"

def current_role() -> str:
    return st.session_state.get("role") or "public"

def is_admin_user() -> bool:
    return current_role() == "admin"

def allowed_pages_for_current_role() -> list[str]:
    return ROLE_PAGES.get(current_role(), ROLE_PAGES["cashier"])

def safe_float(value, default=0.0) -> float:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default

def safe_int(value, default=0) -> int:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default

def normalize_for_compare(value):
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return str(value)
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, dict):
        return {str(k): normalize_for_compare(v) for k, v in sorted(value.items())}
    if isinstance(value, list):
        return [normalize_for_compare(v) for v in value]
    return value

def stable_json(value) -> str:
    return json.dumps(normalize_for_compare(value), sort_keys=True, default=str, ensure_ascii=True)

def make_sku(name: str, category: str = "") -> str:
    base = re.sub(r"[^A-Za-z0-9]+", "-", (name or category or "ITEM").upper()).strip("-")[:18]
    digest = hashlib.sha1(f"{category}|{name}".encode()).hexdigest()[:6].upper()
    return f"{base or 'ITEM'}-{digest}"

def record_fingerprint(doc: dict) -> str:
    ignored = {
        "_id", "id", "created_at", "updated_at", "restored_at", "deleted_at",
        "deleted_by", "deleted_reason", "version", "dedupe_key", "restore_source_id",
        "payment_history", "reminder_history", "return_history",
    }
    significant = {k: normalize_for_compare(v) for k, v in doc.items() if k not in ignored}
    return hashlib.sha256(stable_json(significant).encode()).hexdigest()

def line_item_total(item: dict) -> float:
    qty = safe_float(item.get("quantity"), 1.0)
    unit_price = safe_float(item.get("unit_price", item.get("selling_price")), 0.0)
    discount = safe_float(item.get("discount_amount"), 0.0)
    return max(round((unit_price * qty) - discount, 2), 0.0)

def line_item_profit(item: dict) -> float:
    qty = safe_float(item.get("quantity"), 1.0)
    total = line_item_total(item)
    cost = safe_float(item.get("cost_price", item.get("buying_price")), 0.0)
    return round(total - (cost * qty), 2)

def doc_total_amount(doc: dict) -> float:
    items = doc.get("line_items")
    if isinstance(items, list) and items:
        return round(sum(line_item_total(item) for item in items), 2)
    qty = safe_float(doc.get("quantity"), 1.0) or 1.0
    gross = safe_float(doc.get("gross_amount"), safe_float(doc.get("selling_price"), 0.0) * qty)
    discount = safe_float(doc.get("discount_amount"), 0.0)
    return max(round(gross - discount, 2), 0.0)

def doc_profit_amount(doc: dict) -> float:
    items = doc.get("line_items")
    if isinstance(items, list) and items:
        return round(sum(line_item_profit(item) for item in items), 2)
    qty = safe_float(doc.get("quantity"), 1.0) or 1.0
    return round(doc_total_amount(doc) - (safe_float(doc.get("buying_price"), 0.0) * qty), 2)

def prepare_transaction_doc(doc: dict) -> dict:
    now = str(datetime.now())
    doc.setdefault("id", get_next_id())
    doc.setdefault("created_at", now)
    doc.setdefault("created_by", current_user())
    doc.setdefault("recorded_by", current_user())
    doc.setdefault("version", 1)
    doc.setdefault("transaction_type", "sale")
    doc.setdefault("quantity", 1)
    doc.setdefault("discount_amount", 0.0)
    doc.setdefault("gross_amount", round(safe_float(doc.get("selling_price")) * safe_float(doc.get("quantity"), 1), 2))
    doc["dedupe_key"] = doc.get("dedupe_key") or record_fingerprint(doc)
    return doc

def audit_log(record_id, action: str, changes: list[dict], collection: str = "sales"):
    get_audit_col().insert_one({
        "record_id": record_id,
        "collection": collection,
        "action": action,
        "changes": changes,
        "user": current_user(),
        "role": current_role(),
        "timestamp": str(datetime.now()),
    })

def build_changes(old_doc: dict, set_fields: dict | None = None, push_fields: dict | None = None, unset_fields: list[str] | None = None) -> list[dict]:
    changes = []
    for field, new_value in (set_fields or {}).items():
        old_value = old_doc.get(field)
        if normalize_for_compare(old_value) != normalize_for_compare(new_value):
            changes.append({"field": field, "old_value": normalize_for_compare(old_value), "new_value": normalize_for_compare(new_value)})
    for field, new_value in (push_fields or {}).items():
        changes.append({"field": field, "old_value": normalize_for_compare(old_doc.get(field, [])), "new_value": normalize_for_compare(new_value)})
    for field in (unset_fields or []):
        if field in old_doc:
            changes.append({"field": field, "old_value": normalize_for_compare(old_doc.get(field)), "new_value": None})
    return changes

def update_record_with_audit(
    record_id,
    set_fields: dict | None = None,
    action: str = "update",
    expected_version: int | None = None,
    push_fields: dict | None = None,
    unset_fields: list[str] | None = None,
    include_deleted: bool = False,
) -> tuple[bool, str]:
    try:
        rid = int(record_id)
    except (TypeError, ValueError):
        rid = record_id
    current = get_col().find_one(active_filter({"id": rid}, include_deleted=include_deleted))
    if not current:
        return False, "missing"

    current_version = safe_int(current.get("version"), 1)
    if expected_version is not None and current_version != safe_int(expected_version, 1):
        return False, "conflict"

    changed_fields = dict(set_fields or {})
    changed_fields["updated_at"] = str(datetime.now())
    changed_fields["updated_by"] = current_user()
    changed_fields["version"] = current_version + 1
    changes = build_changes(current, set_fields, push_fields, unset_fields)

    update = {"$set": changed_fields}
    if push_fields:
        update["$push"] = push_fields
    if unset_fields:
        update["$unset"] = {field: "" for field in unset_fields}

    version_filter = {"_id": current["_id"]}
    if "version" in current:
        version_filter["version"] = current_version
    else:
        version_filter["version"] = {"$exists": False}

    result = get_col().update_one(version_filter, update)
    if result.matched_count == 0:
        return False, "conflict"
    if changes:
        audit_log(rid, action, changes)
    return True, "updated" if result.modified_count else "noop"

def insert_transaction(doc: dict, action: str = "create") -> int:
    doc = prepare_transaction_doc(doc)
    get_col().insert_one(doc)
    audit_log(doc.get("id"), action, [{"field": "record", "old_value": None, "new_value": normalize_for_compare({k: v for k, v in doc.items() if k != "_id"})}])
    return doc["id"]

def paginated_slice(df: pd.DataFrame, key: str, default_size: int = 25) -> pd.DataFrame:
    if df.empty:
        return df
    size = st.selectbox("Rows per page", [10, 25, 50, 100], index=[10, 25, 50, 100].index(default_size) if default_size in [10, 25, 50, 100] else 1, key=f"{key}_size")
    pages = max(1, math.ceil(len(df) / size))
    page = st.number_input("Page", min_value=1, max_value=pages, value=min(st.session_state.get(f"{key}_page", 1), pages), step=1, key=f"{key}_page")
    start = (int(page) - 1) * size
    st.caption(f"Showing {start + 1}-{min(start + size, len(df))} of {len(df)} records")
    return df.iloc[start:start + size]

def loyalty_for_spend(spend: float) -> dict:
    tier = LOYALTY_TIERS[0]
    for candidate in LOYALTY_TIERS:
        if spend >= candidate["min_spend"]:
            tier = candidate
    return tier

def customer_spend(customer_name: str) -> float:
    if not customer_name:
        return 0.0
    docs = list(get_col().find(
        active_filter({"customer_name": {"$regex": f"^{re.escape(customer_name.strip())}$", "$options": "i"}, "transaction_type": "sale"}),
        {"_id": 0, "selling_price": 1, "quantity": 1, "discount_amount": 1, "gross_amount": 1, "line_items": 1},
    ))
    return sum(doc_total_amount(d) for d in docs)

# =====================================================
# DATA HELPERS
# =====================================================

@st.cache_data(ttl=30)
def fetch_all() -> pd.DataFrame:
    docs = list(get_col().find(active_filter(), {"_id": 0}))
    if not docs:
        return pd.DataFrame()
    df = pd.DataFrame(docs)
    if "quantity" not in df.columns:
        df["quantity"] = 1
    for c in ["buying_price", "selling_price", "amount_paid", "pending_amount", "quantity", "discount_amount", "gross_amount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["quantity"] = df["quantity"].replace(0, 1)
    for c in ["payment_received", "delay_status", "version"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    if "sale_date" in df.columns:
        df["sale_date"] = pd.to_datetime(df["sale_date"], errors="coerce")
    for col in ["vendor", "product_category", "product_description", "notes", "customer_phone", "transaction_type", "return_reason", "expense_category", "recorded_by", "sku", "coupon_code", "discount_reason"]:
        if col not in df.columns:
            df[col] = ""
    pending_only = is_pending_payment_record(df)
    df.loc[pending_only, "transaction_type"] = "pending_payment"
    df["transaction_type"] = df["transaction_type"].replace("", "sale").fillna("sale")
    expense_only = df["transaction_type"].eq("expense")
    return_only = df["transaction_type"].eq("return")
    sale_only = df["transaction_type"].eq("sale")
    df["total_amount"] = df.apply(lambda row: doc_total_amount(row.to_dict()), axis=1)
    df["profit"] = df.apply(lambda row: doc_profit_amount(row.to_dict()), axis=1)
    df.loc[pending_only, "profit"] = 0.0
    df.loc[expense_only, "profit"] = -df.loc[expense_only, "selling_price"]
    df.loc[return_only, "profit"] = -df.loc[return_only, "amount_paid"].clip(lower=0)
    df.loc[expense_only, "total_amount"] = 0.0
    df.loc[return_only, "total_amount"] = -df.loc[return_only, "amount_paid"].clip(lower=0)
    df["margin"] = (df["profit"] / df["total_amount"].replace(0, 1) * 100).round(2)
    df.loc[pending_only | expense_only | return_only | ~sale_only, "margin"] = 0.0
    return df

def invalidate_cache():
    fetch_all.clear()

def is_pending_payment_record(df: pd.DataFrame) -> pd.Series:
    if df.empty:
        return pd.Series(dtype=bool)
    transaction_type = df["transaction_type"] if "transaction_type" in df.columns else pd.Series([""] * len(df), index=df.index)
    description = df["product_description"] if "product_description" in df.columns else pd.Series([""] * len(df), index=df.index)
    category = df["product_category"] if "product_category" in df.columns else pd.Series([""] * len(df), index=df.index)
    buying = df["buying_price"] if "buying_price" in df.columns else pd.Series([0] * len(df), index=df.index)
    pending = df["pending_amount"] if "pending_amount" in df.columns else pd.Series([0] * len(df), index=df.index)
    return (
        transaction_type.eq("pending_payment")
        | (
            description.astype(str).str.lower().eq("pending payment")
            & category.astype(str).eq("Other")
            & buying.eq(0)
            & pending.gt(0)
        )
    )

def accounted_sales(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    t = df["transaction_type"] if "transaction_type" in df.columns else pd.Series(["sale"] * len(df), index=df.index)
    return df[t.eq("sale") & ~is_pending_payment_record(df)].copy()

def expense_records(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "transaction_type" not in df.columns:
        return df.iloc[0:0].copy() if not df.empty else df
    return df[df["transaction_type"].eq("expense")].copy()

def return_records(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "transaction_type" not in df.columns:
        return df.iloc[0:0].copy() if not df.empty else df
    return df[df["transaction_type"].eq("return")].copy()

def metrics(df: pd.DataFrame) -> dict:
    if df.empty:
        return dict(sales=0, revenue=0, collected=0, profit=0, expenses=0, net_profit=0, pending=0, delayed=0, margin=0, customers=0)
    sales_df = accounted_sales(df)
    expenses_df = expense_records(df)
    returns_df = return_records(df)
    revenue = sales_df["total_amount"].sum() if "total_amount" in sales_df.columns else sales_df["selling_price"].sum()
    collected = sales_df["amount_paid"].sum() if "amount_paid" in sales_df.columns else 0
    returns_total = returns_df["amount_paid"].sum() if "amount_paid" in returns_df.columns else 0
    expense_total = expenses_df["selling_price"].sum() if "selling_price" in expenses_df.columns else 0
    gross_profit = sales_df["profit"].sum() - returns_total
    net_profit = gross_profit - expense_total
    return dict(
        sales     = len(sales_df),
        revenue   = revenue,
        collected = max(collected - returns_total, 0),
        profit    = gross_profit,
        expenses  = expense_total,
        net_profit= net_profit,
        pending   = df["pending_amount"].sum(),
        delayed   = int((sales_df["delay_status"] == 1).sum()) if not sales_df.empty else 0,
        margin    = sales_df["margin"].mean() if not sales_df.empty else 0,
        customers = sales_df["customer_name"].nunique() if not sales_df.empty else 0,
    )

def to_excel(df: pd.DataFrame) -> BytesIO:
    out = BytesIO()
    ex = df.copy()
    if "sale_date" in ex.columns:
        ex["sale_date"] = ex["sale_date"].astype(str)
    qty = ex["quantity"] if "quantity" in ex.columns else pd.Series([1] * len(ex), index=ex.index)
    if "discount_amount" not in ex.columns:
        ex["discount_amount"] = 0.0
    if "gross_amount" not in ex.columns:
        ex["gross_amount"] = ex["selling_price"] * qty
    ex["total_amount"] = ex.apply(lambda row: doc_total_amount(row.to_dict()), axis=1)
    ex["profit"]       = ex.apply(lambda row: doc_profit_amount(row.to_dict()), axis=1)
    pending_only = is_pending_payment_record(ex)
    t_col = ex["transaction_type"] if "transaction_type" in ex.columns else pd.Series(["sale"] * len(ex), index=ex.index)
    ex.loc[pending_only, "profit"] = 0.0
    ex.loc[t_col.eq("expense"), "profit"] = -ex.loc[t_col.eq("expense"), "selling_price"]
    ex.loc[t_col.eq("return"), "profit"] = -ex.loc[t_col.eq("return"), "amount_paid"].clip(lower=0)
    ex.loc[t_col.eq("expense"), "total_amount"] = 0.0
    ex.loc[t_col.eq("return"), "total_amount"] = -ex.loc[t_col.eq("return"), "amount_paid"].clip(lower=0)
    ex["profit_margin"] = (ex["profit"] / ex["total_amount"].replace(0, 1) * 100).round(2)
    ex.loc[pending_only | t_col.isin(["expense", "return"]), "profit_margin"] = 0.0
    ex["status"]  = ex["payment_received"].map({0: "Pending", 1: "Received"})
    ex["delayed"] = ex["delay_status"].map({0: "No", 1: "Yes"})
    ordered = ["id","transaction_type","customer_name","customer_phone","sale_date","vendor","sku","product_category","expense_category","product_description","quantity","buying_price","selling_price","gross_amount","discount_amount","coupon_code","discount_reason","total_amount","profit","profit_margin","amount_paid","pending_amount","status","delayed","payment_method","notes","recorded_by","created_at"]
    cols = [c for c in ordered if c in ex.columns]
    ex = ex[cols]
    ex.columns = [c.replace("_", " ").title() for c in ex.columns]
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        ex.to_excel(w, index=False)
        ws = w.sheets["Sheet1"]
        for i, col in enumerate(ex.columns, 1):
            ml = max(ex.iloc[:, i-1].astype(str).str.len().max(), len(col)) + 4
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(ml, 45)
        blue_fill = PatternFill("solid", fgColor="2E6FD8")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="E8EEF8")
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")
    out.seek(0)
    return out

def get_existing_customers(include_phone: bool = False):
    pipeline = [
        {"$match": active_filter({"customer_name": {"$ne": None, "$ne": ""}})},
        {"$group": {
            "_id": "$customer_name",
            **({"phone": {"$first": "$customer_phone"}} if include_phone else {}),
            "visits": {"$sum": 1},
            "last_sale": {"$max": "$sale_date"},
        }},
        {"$sort": {"_id": 1}},
    ]
    return list(get_col().aggregate(pipeline))

def get_existing_customers_with_phone():
    return get_existing_customers(include_phone=True)

# =====================================================
# HELPERS
# =====================================================

def page_header(title, sub):
    st.markdown(f"<div class='page-title'>{html_escape(title)}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='page-sub'>{html_escape(sub)}</div>", unsafe_allow_html=True)

def sec(label):
    st.markdown(f"<div class='sec-head'>{html_escape(label)}</div>", unsafe_allow_html=True)

def rule():
    st.markdown("<hr class='rule'>", unsafe_allow_html=True)

def rule_sm():
    st.markdown("<hr class='rule-sm'>", unsafe_allow_html=True)

def is_admin():
    return st.session_state.get("logged_in", False)

def logout_user(message: str | None = None):
    st.session_state.logged_in = False
    st.session_state.username = None
    st.session_state.role = None
    st.session_state.staff_id = None
    st.session_state.last_activity = None
    if message:
        st.warning(message)

def enforce_session_timeout() -> bool:
    if not st.session_state.get("logged_in", False):
        return True
    last_activity = st.session_state.get("last_activity", time.time())
    if time.time() - last_activity > SESSION_TIMEOUT_SECONDS:
        logout_user("Your admin session expired after inactivity. Please sign in again.")
        return False
    st.session_state.last_activity = time.time()
    return True

def is_valid_indian_phone(phone: str) -> bool:
    phone = (phone or "").strip()
    if not phone:
        return True
    return bool(re.fullmatch(r"(?:\+91[\s-]?|91[\s-]?|0)?[6-9]\d{9}", phone))

def normalize_phone(phone: str) -> str:
    digits = re.sub(r"\D", "", phone or "")
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits[-10:] if len(digits) >= 10 else digits

def duplicate_sale_exists(
    customer_name: str = "",
    sale_date=None,
    selling_price: float = 0.0,
    exclude_id=None,
    quantity=1,
    description="",
    category="",
    idempotency_key: str | None = None,
    dedupe_key: str | None = None,
) -> bool:
    if dedupe_key:
        query = active_filter({"dedupe_key": dedupe_key})
    elif idempotency_key:
        query = active_filter({"idempotency_key": idempotency_key})
    else:
        return False
    if exclude_id is not None:
        query["id"] = {"$ne": exclude_id}
    return get_col().count_documents(query, limit=1) > 0

def duplicate_pending_exists(customer_name: str = "", pending_date=None, amount: float = 0.0, exclude_id=None, idempotency_key: str | None = None, dedupe_key: str | None = None) -> bool:
    if dedupe_key:
        query = active_filter({"dedupe_key": dedupe_key, "transaction_type": "pending_payment"})
    elif idempotency_key:
        query = active_filter({"idempotency_key": idempotency_key, "transaction_type": "pending_payment"})
    else:
        return False
    if exclude_id is not None:
        query["id"] = {"$ne": exclude_id}
    return get_col().count_documents(query, limit=1) > 0

def append_payment_history(row_id, amount: float, method: str = "", note: str = ""):
    update_record_with_audit(
        row_id,
        action="payment_history_append",
        push_fields={"payment_history": {
            "amount": round(float(amount), 2),
            "method": method,
            "note": note,
            "paid_at": str(datetime.now()),
            "recorded_by": current_user(),
        }},
    )

def get_inventory_items(active_only: bool = True) -> list[dict]:
    query = active_filter() if active_only else {}
    return list(get_db()["inventory"].find(query, {"_id": 0}).sort([("category", 1), ("name", 1)]))

def inventory_option_label(item: dict) -> str:
    sku = str(item.get("sku", "") or "No SKU")
    name = str(item.get("name", "") or "Unnamed")
    qty = safe_int(item.get("quantity"), 0)
    return f"{sku} — {name} ({qty} in stock)"

def increment_inventory(category: str, quantity: int, item_name: str = "", vendor: str = "", cost_price: float = 0.0, sku: str = "") -> bool:
    if quantity <= 0:
        return False
    item_name = (item_name or category or "Unspecified").strip()
    sku = (sku or make_sku(item_name, category)).strip()
    key = {"sku": sku}
    update = {
        "$inc": {"quantity": int(quantity)},
        "$set": {"category": category, "name": item_name, "sku": sku, "vendor": vendor.strip(), "updated_at": str(datetime.now())},
        "$setOnInsert": {"min_stock": 0, "cost_price": round(float(cost_price or 0), 2), "sell_price": 0.0, "notes": "Auto-created from transaction", "created_at": str(datetime.now())},
    }
    get_db()["inventory"].update_one(key, update, upsert=True)
    return True

def decrement_inventory_for_sale(category: str, quantity: int, item_name: str = "", sku: str = "") -> bool:
    if quantity <= 0:
        return False
    inv_col = get_db()["inventory"]
    candidates = []
    if sku:
        candidates.append(active_filter({"sku": {"$regex": f"^{re.escape(sku.strip())}$", "$options": "i"}, "quantity": {"$gte": int(quantity)}}))
    if item_name:
        candidates.append(active_filter({"category": category, "name": {"$regex": f"^{re.escape(item_name.strip())}$", "$options": "i"}, "quantity": {"$gte": int(quantity)}}))
    for query in candidates:
        result = inv_col.update_one(query, {"$inc": {"quantity": -int(quantity)}, "$set": {"updated_at": str(datetime.now())}})
        if result.modified_count > 0:
            return True
    return False

def restore_inventory_for_transaction(row: dict | pd.Series) -> bool:
    t = row.get("transaction_type", "sale")
    if t not in ("sale", "return"):
        return False
    if t == "return":
        return decrement_inventory_for_sale(
            row.get("product_category", "Other"),
            safe_int(row.get("quantity"), 1),
            row.get("product_description", ""),
            row.get("sku", ""),
        )
    items = row.get("line_items")
    if isinstance(items, list) and items:
        restored = False
        for item in items:
            restored = increment_inventory(
                item.get("category", "Other"),
                safe_int(item.get("quantity"), 1),
                item.get("description", "") or item.get("category", "Other"),
                row.get("vendor", ""),
                safe_float(item.get("cost_price"), 0.0),
                item.get("sku", ""),
            ) or restored
        return restored
    return increment_inventory(
        row.get("product_category", "Other"),
        int(row.get("quantity", 1) or 1),
        row.get("product_description", "") or row.get("product_category", "Other"),
        row.get("vendor", ""),
        float(row.get("buying_price", 0) or 0),
        row.get("sku", ""),
    )

def log_reminder(row_id, channel="WhatsApp"):
    update_record_with_audit(
        row_id,
        action="log_reminder",
        push_fields={"reminder_history": {"channel": channel, "sent_at": str(datetime.now()), "logged_by": current_user()}},
    )

def whatsapp_link(phone: str, customer: str, amount: float) -> str | None:
    digits = normalize_phone(phone)
    if not digits or len(digits) != 10:
        return None
    text = f"Hello {customer}, this is a reminder from Vinay Boutique. Pending amount: Rs {amount:,.2f}. Please share payment update when convenient."
    from urllib.parse import quote
    return f"https://wa.me/91{digits}?text={quote(text)}"

def apply_payment_update(row_id, amount: float, method: str, note: str, expected_version: int | None = None) -> tuple[bool, str]:
    current = get_col().find_one(active_filter({"id": int(row_id)}))
    if not current:
        return False, "missing"
    amount = round(float(amount), 2)
    if amount <= 0:
        return False, "amount"
    total_due = doc_total_amount(current)
    current_paid = safe_float(current.get("amount_paid"), 0.0)
    current_pending = safe_float(current.get("pending_amount"), max(total_due - current_paid, 0.0))
    pay_now = min(amount, current_pending)
    new_paid = min(round(current_paid + pay_now, 2), total_due)
    new_pending = max(round(total_due - new_paid, 2), 0.0)
    history = {
        "amount": pay_now,
        "method": method,
        "note": note,
        "paid_at": str(datetime.now()),
        "recorded_by": current_user(),
    }
    return update_record_with_audit(
        row_id,
        set_fields={
            "payment_received": 1 if new_pending == 0 else 0,
            "amount_paid": new_paid,
            "pending_amount": new_pending,
            "payment_method": method,
        },
        push_fields={"payment_history": history},
        action="payment_update",
        expected_version=expected_version,
    )

def mark_transaction_paid(row_id, method: str = "", note: str = "Marked paid", expected_version: int | None = None) -> tuple[bool, str]:
    current = get_col().find_one(active_filter({"id": int(row_id)}))
    if not current:
        return False, "missing"
    total_due = doc_total_amount(current)
    delta = max(round(total_due - safe_float(current.get("amount_paid"), 0.0), 2), 0.0)
    if delta <= 0:
        return update_record_with_audit(
            row_id,
            set_fields={"payment_received": 1, "pending_amount": 0.0},
            action="mark_paid",
            expected_version=expected_version,
        )
    return apply_payment_update(row_id, delta, method or current.get("payment_method", ""), note, expected_version=expected_version)

def soft_delete_transaction(row_id, reason: str = "", expected_version: int | None = None) -> tuple[bool, str]:
    current = get_col().find_one(active_filter({"id": int(row_id)}))
    if not current:
        return False, "missing"
    ok, status = update_record_with_audit(
        row_id,
        set_fields={
            "deleted_at": str(datetime.now()),
            "deleted_by": current_user(),
            "deleted_reason": reason.strip()[:300],
        },
        action="soft_delete",
        expected_version=expected_version if expected_version is not None else safe_int(current.get("version"), 1),
    )
    if ok and current.get("transaction_type", "sale") in ("sale", "return"):
        restore_inventory_for_transaction(current)
    return ok, status

def recover_deleted_transaction(row_id) -> tuple[bool, str]:
    current = get_col().find_one(active_filter({"id": int(row_id)}, include_deleted=True))
    if not current or "deleted_at" not in current:
        return False, "missing"
    if current.get("transaction_type", "sale") == "return":
        increment_inventory(
            current.get("product_category", "Other"),
            safe_int(current.get("quantity"), 1),
            current.get("product_description", ""),
            current.get("vendor", ""),
            safe_float(current.get("buying_price"), 0.0),
            current.get("sku", ""),
        )
    elif current.get("transaction_type", "sale") == "sale":
        items = current.get("line_items")
        if isinstance(items, list) and items:
            failed = [
                item for item in items
                if not decrement_inventory_for_sale(item.get("category", "Other"), safe_int(item.get("quantity"), 1), item.get("description", ""), item.get("sku", ""))
            ]
            if failed:
                return False, "inventory"
        elif not decrement_inventory_for_sale(
            current.get("product_category", "Other"),
            safe_int(current.get("quantity"), 1),
            current.get("product_description", ""),
            current.get("sku", ""),
        ):
            return False, "inventory"
    return update_record_with_audit(
        row_id,
        set_fields={"recovered_at": str(datetime.now()), "recovered_by": current_user()},
        unset_fields=["deleted_at", "deleted_by", "deleted_reason"],
        action="recover",
        expected_version=safe_int(current.get("version"), 1),
        include_deleted=True,
    )

# =====================================================
# ADD SALE PAGE
# =====================================================

def customer_autocomplete(label: str, customers, key: str, placeholder: str = "") -> str:
    names = [str(r.get("_id", "")).strip() for r in customers if str(r.get("_id", "")).strip()]
    try:
        return st.selectbox(
            label,
            options=names,
            index=None,
            placeholder=placeholder or "Start typing a customer name",
            accept_new_options=True,
            key=key,
        ) or ""
    except TypeError:
        typed = st.text_input(label, placeholder=placeholder or "Customer name", key=f"{key}_text")
        if typed:
            matches = [n for n in names if typed.lower() in n.lower()][:5]
            if matches:
                picked = st.selectbox("Suggestions", ["Use typed name"] + matches, key=f"{key}_suggestions")
                return typed if picked == "Use typed name" else picked
        return typed

def render_new_sale_form(public=False, show_header=True):
    if public:
        st.markdown("""
        <div class='pub-banner'>
            <div class='pub-banner-title'>Vinay Boutique</div>
            <div class='pub-banner-sub'>◆ Record a New Sale</div>
        </div>
        """, unsafe_allow_html=True)
    elif show_header:
        page_header("New Sale", "Record a Transaction")

    if public:
        ctype = "New Customer"
        st.caption("Public entry accepts new sale details only. Existing customer lookup is available after admin sign-in.")
    else:
        ctype = st.radio("", ["New Customer", "Existing Customer"], horizontal=True)
    rule_sm()

    cname, cphone = "", ""

    if ctype == "Existing Customer":
        existing = get_existing_customers(include_phone=True)

        if existing:
            opts  = [f"{r['_id']}  —  {r.get('phone','') or 'No phone'}" for r in existing]
            sel   = st.selectbox("Select Customer", opts)
            cname = sel.split("  —  ")[0].strip()
            rec   = next((r for r in existing if r["_id"] == cname), {})
            cphone = rec.get("phone", "")
            ca, cb, cc = st.columns(3)
            ca.info(f"**Name:** {cname}")
            cb.info(f"**Phone:** {cphone or 'N/A'}")
            cc.info(f"**Visits:** {rec.get('visits', '—')}")
        else:
            st.warning("No existing customers found.")
            ctype = "New Customer"

    sale_token_key = "public_sale_token" if public else "admin_sale_token"
    if sale_token_key not in st.session_state:
        st.session_state[sale_token_key] = str(uuid.uuid4())
    inventory_items = [] if public else get_inventory_items()
    inventory_by_sku = {str(item.get("sku", "")): item for item in inventory_items if item.get("sku")}

    with st.form("sale_form", clear_on_submit=True):
        sec("Customer")
        c1, c2, c3 = st.columns(3)
        with c1:
            cname  = st.text_input("Customer Name *", value=cname,
                                   disabled=(ctype == "Existing Customer"))
        with c2:
            if public and ctype == "Existing Customer":
                cphone = ""
                st.text_input("Phone", value="", placeholder="(Admin access required)", disabled=True)
            else:
                cphone = st.text_input("Phone", value=cphone, placeholder="+91 XXXXXXXXXX",
                                       disabled=(ctype == "Existing Customer"))
        with c3:
            sdate = st.date_input("Sale Date", date.today())

        sec("Product")
        selected_inventory = None
        sku = ""
        if not public and inventory_items:
            inv_choice = st.selectbox(
                "Inventory SKU",
                ["Manual entry"] + [str(item.get("sku", "")) for item in inventory_items if item.get("sku")],
                format_func=lambda x: x if x == "Manual entry" else inventory_option_label(inventory_by_sku.get(x, {})),
            )
            if inv_choice != "Manual entry":
                sku = inv_choice
                selected_inventory = inventory_by_sku.get(inv_choice)
        p1, p2, p3 = st.columns(3)
        default_cat = selected_inventory.get("category", CATEGORIES[0]) if selected_inventory else CATEGORIES[0]
        with p1:
            cat  = st.selectbox("Category *", CATEGORIES, index=CATEGORIES.index(default_cat) if default_cat in CATEGORIES else 0)
        with p2:
            vend = st.text_input("Vendor / Supplier", value=str(selected_inventory.get("vendor", "")) if selected_inventory else "")
        with p3:
            qty  = st.number_input("Quantity", min_value=1, step=1, value=1)
        desc = st.text_area("Description", value=str(selected_inventory.get("name", "")) if selected_inventory else "", placeholder="Fabric, colour, design details…", height=70)

        sec("Pricing & Payment")
        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1:
            buy = st.number_input("Buying Price (₹) *", min_value=0.0, value=safe_float(selected_inventory.get("cost_price"), 0.0) if selected_inventory else 0.0, step=100.0, format="%.2f")
        with pr2:
            sell = st.number_input("Selling Price (₹) *", min_value=0.0, value=safe_float(selected_inventory.get("sell_price"), 0.0) if selected_inventory else 0.0, step=100.0, format="%.2f")
        with pr3:
            coupon_code = st.text_input("Coupon Code").strip().upper()
        with pr4:
            discount_reason = st.text_input("Discount Reason", placeholder="Optional")

        gross_amt = round(sell * qty, 2)
        loyalty = loyalty_for_spend(customer_spend(cname.strip())) if (not public and cname.strip()) else LOYALTY_TIERS[0]
        coupon_pct = safe_float(COUPONS.get(coupon_code, {}).get("discount_pct"), 0.0) if coupon_code else 0.0
        suggested_discount_pct = max(coupon_pct, safe_float(loyalty.get("discount_pct"), 0.0))
        if not public and cname.strip():
            st.caption(f"Loyalty: {loyalty['tier']} · suggested discount {loyalty['discount_pct']:.0f}% · {loyalty['points_rate']} point(s) per ₹100")

        d1, d2, d3, d4 = st.columns(4)
        with d1:
            discount_mode = st.selectbox("Discount Type", ["None", "Amount", "Percent", "Coupon/Loyalty"])
        with d2:
            discount_value = st.number_input("Discount Value", min_value=0.0, step=50.0, format="%.2f")
        with d3:
            split_payment = st.checkbox("Split Payment")
        with d4:
            split_count = st.number_input("Payment Lines", min_value=2, max_value=4, value=2, step=1, disabled=not split_payment)

        if discount_mode == "Percent":
            discount_amt = round(gross_amt * min(discount_value, 100.0) / 100, 2)
        elif discount_mode == "Coupon/Loyalty":
            discount_amt = round(gross_amt * suggested_discount_pct / 100, 2)
        elif discount_mode == "Amount":
            discount_amt = round(discount_value, 2)
        else:
            discount_amt = 0.0
        discount_amt = min(discount_amt, gross_amt)
        total_amt = max(round(gross_amt - discount_amt, 2), 0.0)

        payment_splits = []
        if split_payment:
            for i in range(int(split_count)):
                pc1, pc2 = st.columns(2)
                with pc1:
                    split_method = st.selectbox("Method", PAYMENT_METHODS, key=f"sale_split_method_{i}")
                with pc2:
                    split_amount = st.number_input("Amount (₹)", min_value=0.0, step=100.0, format="%.2f", key=f"sale_split_amount_{i}")
                if split_amount > 0:
                    payment_splits.append({"method": split_method, "amount": round(float(split_amount), 2)})
            paid_amt = round(sum(p["amount"] for p in payment_splits), 2)
            pm = "Split"
        else:
            pr5, pr6 = st.columns(2)
            with pr5:
                paid_amt = st.number_input("Amount Paid (₹)", min_value=0.0, step=100.0, format="%.2f")
            with pr6:
                pm = st.selectbox("Payment Method", PAYMENT_METHODS)

        pending_amt = max(round(total_amt - paid_amt, 2), 0.0)
        profit_amt  = round(total_amt - (buy * qty), 2)
        margin_pct  = round(profit_amt / total_amt * 100, 2) if total_amt > 0 else 0.0

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Pending",        f"₹{pending_amt:,.2f}")
        m2.metric("Profit (Total)", f"₹{profit_amt:,.2f}")
        m3.metric("Margin",         f"{margin_pct:.1f}%")
        m4.metric("Total Value",    f"₹{total_amt:,.2f}")

        notes = st.text_area("Notes", placeholder="Special instructions…", height=60)

        submitted = st.form_submit_button("Save Sale", use_container_width=True)

        if submitted:
            # ── Public form rate limiting ────────────────────────────────
            if public:
                now = time.time()
                window_start = st.session_state.get("pub_window_start", now)
                pub_count    = st.session_state.get("pub_submit_count", 0)
                # Reset counter every 60 seconds
                if now - window_start > 60:
                    st.session_state.pub_window_start  = now
                    st.session_state.pub_submit_count  = 0
                    pub_count = 0
                pub_count += 1
                st.session_state.pub_submit_count = pub_count
                if pub_count > 5:
                    st.error("Too many submissions. Please wait a minute before trying again.")
                    st.stop()

            # ── Input length guards ──────────────────────────────────────
            MAX = {"name": 120, "phone": 20, "vendor": 100, "desc": 500, "notes": 500}
            errs = []
            if len(cname) > MAX["name"]:   errs.append(f"Customer name must be under {MAX['name']} characters.")
            if len(cphone) > MAX["phone"]:  errs.append(f"Phone number must be under {MAX['phone']} characters.")
            if len(vend) > MAX["vendor"]:   errs.append(f"Vendor name must be under {MAX['vendor']} characters.")
            if len(desc) > MAX["desc"]:     errs.append(f"Description must be under {MAX['desc']} characters.")
            if len(notes) > MAX["notes"]:   errs.append(f"Notes must be under {MAX['notes']} characters.")
            if cphone.strip() and not is_valid_indian_phone(cphone):
                errs.append("Enter a valid Indian phone number.")

            if not cname.strip():  errs.append("Customer name is required.")
            if buy  <= 0:          errs.append("Buying price must be > 0.")
            if sell <= 0:          errs.append("Selling price must be > 0.")
            if paid_amt > total_amt: errs.append("Amount paid cannot exceed total sale value.")
            if coupon_code and coupon_code not in COUPONS:
                errs.append("Coupon code was not found.")
            if split_payment and not payment_splits:
                errs.append("Enter at least one split payment amount.")
            if duplicate_sale_exists(idempotency_key=st.session_state[sale_token_key]):
                errs.append("This sale was already submitted. Refresh the form before trying again.")

            if errs:
                for e in errs: st.error(e)
            else:
                if total_amt < buy * qty:
                    st.warning("Selling price is below buying price — this sale will be a loss.")
                doc = {
                    "id":                  get_next_id(),
                    "idempotency_key":     st.session_state[sale_token_key],
                    "customer_name":       cname.strip()[:120],
                    "customer_phone":      cphone.strip()[:20],
                    "sale_date":           str(sdate),
                    "vendor":              vend.strip()[:100],
                    "sku":                 sku,
                    "product_category":    cat,
                    "product_description": desc.strip()[:500],
                    "quantity":            qty,
                    "buying_price":        round(buy, 2),
                    "selling_price":       round(sell, 2),
                    "gross_amount":        gross_amt,
                    "discount_amount":     round(discount_amt, 2),
                    "discount_reason":     discount_reason.strip()[:200],
                    "coupon_code":         coupon_code,
                    "loyalty_tier":        loyalty["tier"],
                    "loyalty_discount_pct": loyalty["discount_pct"],
                    "loyalty_points_earned": int(total_amt // 100) * int(loyalty["points_rate"]),
                    "payment_splits":      payment_splits,
                    "amount_paid":         round(paid_amt, 2),
                    "pending_amount":      pending_amt,
                    "payment_received":    1 if pending_amt == 0 else 0,
                    "delay_status":        0,
                    "payment_method":      pm,
                    "notes":               notes.strip()[:500],
                    "transaction_type":     "sale",
                    "recorded_by":          "Public Form" if public else st.session_state.get("username", "Admin"),
                    "created_at":          str(datetime.now()),
                }
                if paid_amt > 0:
                    doc["payment_history"] = [
                        {
                            "amount": round(float(p.get("amount", 0)), 2),
                            "method": p.get("method", pm),
                            "note": "Initial split payment" if split_payment else "Initial payment",
                            "paid_at": str(datetime.now()),
                            "recorded_by": current_user(),
                        }
                        for p in (payment_splits or [{"amount": paid_amt, "method": pm}])
                    ]
                stock_ok = True
                if not public:
                    stock_ok = decrement_inventory_for_sale(cat, qty, desc.strip(), sku)
                    if not stock_ok:
                        st.warning("Sale saved, but matching inventory stock was not available to deduct.")
                        doc["inventory_warning"] = "No exact SKU/item match found during sale save."
                insert_transaction(doc, action="create_sale")
                st.session_state[sale_token_key] = str(uuid.uuid4())
                invalidate_cache()
                st.success(f"✓ Sale recorded for {cname.strip()}.")
                st.balloons()
                st.rerun()

def render_pending_payment_form():
    sec("Customer")
    customers = get_existing_customers_with_phone()
    if "pending_payment_token" not in st.session_state:
        st.session_state.pending_payment_token = str(uuid.uuid4())

    c1, c2, c3 = st.columns(3)
    with c1:
        pname = customer_autocomplete(
            "Customer Name *",
            customers,
            key="pending_customer_name",
            placeholder="Start typing or enter a new name",
        )
    selected = next((r for r in customers if str(r.get("_id", "")).strip().lower() == pname.strip().lower()), None) if pname else None
    with c2:
        pphone_default = str(selected.get("phone", "") or "") if selected else ""
        phone_key = f"pending_customer_phone_{re.sub(r'[^0-9A-Za-z]+', '_', pphone_default) or 'new'}"
        pphone = st.text_input("Phone", value=pphone_default, placeholder="+91 XXXXXXXXXX", key=phone_key)
    with c3:
        pdate = st.date_input("Date", date.today(), key="pending_payment_date")

    with st.form("pending_payment_form", clear_on_submit=True):
        sec("Outstanding Amount")
        a1, a2, a3 = st.columns([1, 1, 2])
        with a1:
            pending_amount = st.number_input("Pending Amount (₹) *", min_value=0.0, step=100.0, format="%.2f")
        with a2:
            pending_cat = st.selectbox("Category", CATEGORIES, index=CATEGORIES.index("Other"))
        with a3:
            pending_desc = st.text_input("What is it for?", placeholder="e.g. blouse alteration balance, saree advance")
        pnotes = st.text_area("Notes", placeholder="Optional context for this pending payment", height=70)

        linked_sales = []
        if pname:
            linked_sales = list(get_col().find(
                active_filter({"customer_name": {"$regex": f"^{re.escape(pname.strip())}$", "$options": "i"}, "transaction_type": "sale"}),
                {"_id": 0, "id": 1, "sale_date": 1, "product_category": 1, "pending_amount": 1, "selling_price": 1, "quantity": 1},
            ).sort("sale_date", -1).limit(15))
        link_options = ["No link"] + [int(r["id"]) for r in linked_sales if r.get("id") is not None]
        link_sale_id = st.selectbox(
            "Link to existing sale",
            link_options,
            format_func=lambda x: "No link" if x == "No link" else f"#{x} — {next((str(r.get('sale_date','')) + ' · ' + str(r.get('product_category','')) + ' · pending ₹' + str(r.get('pending_amount',0)) for r in linked_sales if r.get('id') == x), '')}",
        )

        submitted = st.form_submit_button("Save Pending Payment", use_container_width=True)

        if submitted:
            MAX = {"name": 120, "phone": 20, "notes": 500}
            errs = []
            if not pname.strip():
                errs.append("Customer name is required.")
            if len(pname) > MAX["name"]:
                errs.append(f"Customer name must be under {MAX['name']} characters.")
            if len(pphone) > MAX["phone"]:
                errs.append(f"Phone number must be under {MAX['phone']} characters.")
            if pphone.strip() and not is_valid_indian_phone(pphone):
                errs.append("Enter a valid Indian phone number.")
            if pending_amount <= 0:
                errs.append("Pending amount must be > 0.")
            if len(pnotes) > MAX["notes"]:
                errs.append(f"Notes must be under {MAX['notes']} characters.")
            if duplicate_pending_exists(idempotency_key=st.session_state.pending_payment_token):
                errs.append("This pending payment was already submitted. Refresh the form before trying again.")

            if errs:
                for e in errs:
                    st.error(e)
            else:
                insert_transaction({
                    "id":                  get_next_id(),
                    "idempotency_key":     st.session_state.pending_payment_token,
                    "customer_name":       pname.strip()[:120],
                    "customer_phone":      pphone.strip()[:20],
                    "sale_date":           str(pdate),
                    "vendor":              "",
                    "product_category":    pending_cat,
                    "product_description": (pending_desc.strip() or "Pending payment")[:500],
                    "quantity":            1,
                    "buying_price":        0.0,
                    "selling_price":       round(pending_amount, 2),
                    "amount_paid":         0.0,
                    "pending_amount":      round(pending_amount, 2),
                    "payment_received":    0,
                    "delay_status":        0,
                    "payment_method":      "Credit",
                    "notes":               pnotes.strip()[:500],
                    "transaction_type":     "pending_payment",
                    "linked_sale_id":       None if link_sale_id == "No link" else int(link_sale_id),
                    "credit_history":       [{"amount": round(pending_amount, 2), "note": "Credit opened", "created_at": str(datetime.now())}],
                    "recorded_by":          st.session_state.get("username", "Admin"),
                    "created_at":          str(datetime.now()),
                }, action="create_pending_payment")
                st.session_state.pending_payment_token = str(uuid.uuid4())
                invalidate_cache()
                st.success(f"✓ Pending payment saved for {pname.strip()} — ₹{pending_amount:,.2f}.")

def render_return_form():
    sec("Return / Refund")
    sales = accounted_sales(fetch_all())
    if sales.empty:
        st.info("No sales available to return.")
        return
    sales = sales.sort_values("sale_date", ascending=False)
    selected_id = st.selectbox(
        "Original Sale",
        sales["id"].tolist(),
        format_func=lambda x: f"#{x} — {sales[sales['id']==x]['customer_name'].values[0]} · {sales[sales['id']==x]['sale_date'].dt.strftime('%d %b %Y').values[0]} · ₹{sales[sales['id']==x]['total_amount'].values[0]:,.0f}",
    )
    row = sales[sales["id"] == selected_id].iloc[0]
    with st.form("return_form", clear_on_submit=True):
        r1, r2, r3 = st.columns(3)
        with r1:
            rdate = st.date_input("Return Date", date.today())
        with r2:
            max_qty = int(row.get("quantity", 1) or 1)
            rqty = st.number_input("Quantity Returned", min_value=1, max_value=max_qty, step=1, value=1)
        with r3:
            refund = st.number_input("Refund Amount (₹)", min_value=0.0, max_value=float(row.get("amount_paid", 0) or 0), value=min(float(row.get("amount_paid", 0) or 0), float(row.get("selling_price", 0) or 0) * 1), step=100.0, format="%.2f")
        method_idx = PAYMENT_METHODS.index(row.get("payment_method")) if row.get("payment_method") in PAYMENT_METHODS else 0
        refund_method = st.selectbox("Refund Method", PAYMENT_METHODS, index=method_idx)
        reason = st.text_area("Reason", placeholder="Fit issue, damaged item, exchange, etc.", height=70)
        submitted = st.form_submit_button("Record Return", use_container_width=True)
        if submitted:
            return_id = get_next_id()
            doc = {
                "id": return_id,
                "transaction_type": "return",
                "original_sale_id": int(selected_id),
                "customer_name": row.get("customer_name", ""),
                "customer_phone": row.get("customer_phone", ""),
                "sale_date": str(rdate),
                "vendor": row.get("vendor", ""),
                "product_category": row.get("product_category", "Other"),
                "product_description": row.get("product_description", ""),
                "quantity": int(rqty),
                "buying_price": float(row.get("buying_price", 0) or 0),
                "selling_price": float(row.get("selling_price", 0) or 0),
                "amount_paid": round(float(refund), 2),
                "pending_amount": 0.0,
                "payment_received": 1,
                "delay_status": 0,
                "payment_method": refund_method,
                "return_reason": reason.strip()[:500],
                "notes": f"Return against sale #{selected_id}",
                "recorded_by": st.session_state.get("username", "Admin"),
                "created_at": str(datetime.now()),
            }
            insert_transaction(doc, action="create_return")
            increment_inventory(doc["product_category"], int(rqty), doc["product_description"], doc["vendor"], doc["buying_price"], row.get("sku", ""))
            update_record_with_audit(
                int(selected_id),
                push_fields={"return_history": {"return_id": return_id, "quantity": int(rqty), "refund": round(float(refund), 2), "created_at": str(datetime.now()), "recorded_by": current_user()}},
                action="return_link",
            )
            invalidate_cache()
            st.success(f"✓ Return recorded against sale #{selected_id}.")
            st.rerun()

def render_expense_form():
    sec("Expense")
    with st.form("expense_form", clear_on_submit=True):
        e1, e2, e3 = st.columns(3)
        with e1:
            edate = st.date_input("Expense Date", date.today())
        with e2:
            ecat = st.selectbox("Expense Category", EXPENSE_CATEGORIES)
        with e3:
            amount = st.number_input("Amount (₹) *", min_value=0.0, step=100.0, format="%.2f")
        v1, v2 = st.columns(2)
        with v1:
            vendor = st.text_input("Paid To / Vendor")
        with v2:
            method = st.selectbox("Payment Method", PAYMENT_METHODS, index=0)
        note = st.text_area("Notes", height=70)
        submitted = st.form_submit_button("Save Expense", use_container_width=True)
        if submitted:
            if amount <= 0:
                st.error("Expense amount must be > 0.")
            else:
                insert_transaction({
                    "id": get_next_id(),
                    "transaction_type": "expense",
                    "customer_name": vendor.strip()[:120] or ecat,
                    "customer_phone": "",
                    "sale_date": str(edate),
                    "vendor": vendor.strip()[:100],
                    "product_category": "Other",
                    "expense_category": ecat,
                    "product_description": ecat,
                    "quantity": 1,
                    "buying_price": 0.0,
                    "selling_price": round(float(amount), 2),
                    "amount_paid": round(float(amount), 2),
                    "pending_amount": 0.0,
                    "payment_received": 1,
                    "delay_status": 0,
                    "payment_method": method,
                    "notes": note.strip()[:500],
                    "recorded_by": st.session_state.get("username", "Admin"),
                    "created_at": str(datetime.now()),
                }, action="create_expense")
                invalidate_cache()
                st.success(f"✓ Expense saved — ₹{amount:,.2f}.")
                st.rerun()

def render_cart_sale_form():
    sec("Customer")
    customers = get_existing_customers(include_phone=True)
    if "cart_sale_token" not in st.session_state:
        st.session_state.cart_sale_token = str(uuid.uuid4())

    c1, c2, c3 = st.columns(3)
    with c1:
        cname = customer_autocomplete("Customer Name *", customers, key="cart_customer_name", placeholder="Start typing or enter a new customer")
    selected = next((r for r in customers if str(r.get("_id", "")).strip().lower() == cname.strip().lower()), None) if cname else None
    with c2:
        cphone = st.text_input("Phone", value=str(selected.get("phone", "") or "") if selected else "", placeholder="+91 XXXXXXXXXX", key="cart_customer_phone")
    with c3:
        sdate = st.date_input("Sale Date", date.today(), key="cart_sale_date")

    inventory_items = get_inventory_items()
    inventory_by_sku = {str(item.get("sku", "")): item for item in inventory_items if item.get("sku")}
    line_count = st.number_input("Items in Cart", min_value=1, max_value=12, value=2, step=1, key="cart_line_count")

    with st.form("cart_sale_form", clear_on_submit=True):
        line_items = []
        for i in range(int(line_count)):
            st.markdown(f"**Item {i + 1}**")
            l1, l2, l3 = st.columns([1.2, 2, 1])
            sku_choice = "Manual entry"
            with l1:
                sku_choice = st.selectbox(
                    "SKU",
                    ["Manual entry"] + [str(item.get("sku", "")) for item in inventory_items if item.get("sku")],
                    format_func=lambda x: x if x == "Manual entry" else inventory_option_label(inventory_by_sku.get(x, {})),
                    key=f"cart_sku_{i}",
                )
            with l2:
                desc = st.text_input("Description", key=f"cart_desc_{i}")
            with l3:
                cat = st.selectbox("Category", CATEGORIES, key=f"cart_cat_{i}")
            l4, l5, l6, l7 = st.columns(4)
            with l4:
                qty = st.number_input("Qty", min_value=1, step=1, value=1, key=f"cart_qty_{i}")
            with l5:
                cost = st.number_input("Cost ₹", min_value=0.0, step=50.0, format="%.2f", key=f"cart_cost_{i}")
            with l6:
                price = st.number_input("Price ₹", min_value=0.0, step=50.0, format="%.2f", key=f"cart_price_{i}")
            with l7:
                item_discount = st.number_input("Discount ₹", min_value=0.0, step=50.0, format="%.2f", key=f"cart_discount_{i}")

            selected_item = inventory_by_sku.get(sku_choice) if sku_choice != "Manual entry" else None
            final_desc = (desc.strip() or (selected_item or {}).get("name", "")).strip()
            final_cat = (selected_item or {}).get("category", cat)
            final_cost = cost or safe_float((selected_item or {}).get("cost_price"), 0.0)
            final_price = price or safe_float((selected_item or {}).get("sell_price"), 0.0)
            if final_desc or final_price > 0:
                line_items.append({
                    "sku": "" if sku_choice == "Manual entry" else sku_choice,
                    "description": final_desc,
                    "category": final_cat,
                    "quantity": int(qty),
                    "cost_price": round(float(final_cost), 2),
                    "unit_price": round(float(final_price), 2),
                    "discount_amount": round(float(item_discount), 2),
                    "line_total": line_item_total({"quantity": qty, "unit_price": final_price, "discount_amount": item_discount}),
                })
        gross_amt = round(sum(safe_float(item["unit_price"]) * safe_float(item["quantity"], 1) for item in line_items), 2)
        total_amt = round(sum(line_item_total(item) for item in line_items), 2)
        profit_amt = round(sum(line_item_profit(item) for item in line_items), 2)
        payment_cols = st.columns(3)
        with payment_cols[0]:
            paid_amt = st.number_input("Amount Paid (₹)", min_value=0.0, step=100.0, format="%.2f", key="cart_paid")
        with payment_cols[1]:
            payment_method = st.selectbox("Payment Method", PAYMENT_METHODS, key="cart_payment_method")
        with payment_cols[2]:
            notes = st.text_input("Notes", key="cart_notes")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Items", len(line_items))
        m2.metric("Bill Total", f"₹{total_amt:,.2f}")
        m3.metric("Profit", f"₹{profit_amt:,.2f}")
        m4.metric("Pending", f"₹{max(total_amt - paid_amt, 0):,.2f}")

        submitted = st.form_submit_button("Save Cart Sale", use_container_width=True)
        if submitted:
            errs = []
            if not cname.strip():
                errs.append("Customer name is required.")
            if cphone.strip() and not is_valid_indian_phone(cphone):
                errs.append("Enter a valid Indian phone number.")
            if not line_items:
                errs.append("Add at least one cart item.")
            if any(not item.get("description") for item in line_items):
                errs.append("Each cart line needs a description or SKU.")
            if any(safe_float(item.get("unit_price")) <= 0 for item in line_items):
                errs.append("Each cart line needs a selling price.")
            if paid_amt > total_amt:
                errs.append("Amount paid cannot exceed bill total.")
            if duplicate_sale_exists(idempotency_key=st.session_state.cart_sale_token):
                errs.append("This cart sale was already submitted. Refresh the form before trying again.")

            if errs:
                for err in errs:
                    st.error(err)
            else:
                pending_amt = max(round(total_amt - paid_amt, 2), 0.0)
                doc = {
                    "id": get_next_id(),
                    "idempotency_key": st.session_state.cart_sale_token,
                    "transaction_type": "sale",
                    "sale_mode": "cart",
                    "customer_name": cname.strip()[:120],
                    "customer_phone": cphone.strip()[:20],
                    "sale_date": str(sdate),
                    "vendor": "",
                    "sku": "",
                    "product_category": "Mixed",
                    "product_description": f"Cart sale ({len(line_items)} items)",
                    "line_items": line_items,
                    "quantity": sum(safe_int(item.get("quantity"), 1) for item in line_items),
                    "buying_price": 0.0,
                    "selling_price": 0.0,
                    "gross_amount": gross_amt,
                    "discount_amount": round(gross_amt - total_amt, 2),
                    "amount_paid": round(float(paid_amt), 2),
                    "pending_amount": pending_amt,
                    "payment_received": 1 if pending_amt == 0 else 0,
                    "delay_status": 0,
                    "payment_method": payment_method,
                    "payment_splits": [{"method": payment_method, "amount": round(float(paid_amt), 2)}] if paid_amt > 0 else [],
                    "notes": notes.strip()[:500],
                    "recorded_by": current_user(),
                    "created_at": str(datetime.now()),
                }
                if paid_amt > 0:
                    doc["payment_history"] = [{
                        "amount": round(float(paid_amt), 2),
                        "method": payment_method,
                        "note": "Initial cart payment",
                        "paid_at": str(datetime.now()),
                        "recorded_by": current_user(),
                    }]
                missing_stock = []
                for item in line_items:
                    if not decrement_inventory_for_sale(item.get("category", "Other"), safe_int(item.get("quantity"), 1), item.get("description", ""), item.get("sku", "")):
                        missing_stock.append(item.get("sku") or item.get("description") or "Unknown item")
                if missing_stock:
                    st.warning("Sale saved, but stock was not deducted for: " + ", ".join(missing_stock[:5]))
                    doc["inventory_warning"] = f"No exact inventory match for: {', '.join(missing_stock[:10])}"
                insert_transaction(doc, action="create_cart_sale")
                st.session_state.cart_sale_token = str(uuid.uuid4())
                invalidate_cache()
                st.success(f"✓ Cart sale recorded for {cname.strip()}.")
                st.rerun()

def page_add_sale(public=False):
    if public:
        render_new_sale_form(public=True)
        return

    page_header("Add Sale", "Record a Transaction")
    cart_tab, sale_tab, pending_tab, return_tab, expense_tab = st.tabs(["Cart Sale", "Single Item", "Pending Payment", "Return / Refund", "Expense"])
    with cart_tab:
        render_cart_sale_form()
    with sale_tab:
        render_new_sale_form(public=False, show_header=False)
    with pending_tab:
        render_pending_payment_form()
    with return_tab:
        render_return_form()
    with expense_tab:
        render_expense_form()

# =====================================================
# AUTH HELPERS
# =====================================================

_MAX_ATTEMPTS   = 5
_LOCKOUT_SECS   = 300   # 5-minute lockout after max attempts

def _get_stored_hash() -> bytes | None:
    """
    Return the bcrypt hash of the admin password.

    Priority:
      1. st.secrets["PASSWORD_HASH"]  — a bcrypt hash (preferred for production)
      2. st.secrets["PASSWORD"]       — plain text, hashed on the fly (migration path)
      3. env var PASSWORD_HASH        — bcrypt hash
      4. env var PASSWORD             — plain text, hashed on the fly

    If none of these are set the function returns None and login is blocked.
    """
    try:
        secrets = st.secrets
    except Exception:
        secrets = {}

    # 1. Pre-hashed secret
    h = secrets.get("PASSWORD_HASH") or os.getenv("PASSWORD_HASH", "")
    if h:
        return h.encode() if isinstance(h, str) else h

    # 2. Plain-text secret — hash on the fly (one-time cost per cold start)
    p = secrets.get("PASSWORD") or os.getenv("PASSWORD", "")
    if p:
        return bcrypt.hashpw(p.encode(), bcrypt.gensalt())

    # Nothing configured — fail closed
    return None


def _get_username() -> str | None:
    try:
        u = st.secrets.get("USERNAME") or os.getenv("USERNAME", "")
    except Exception:
        u = os.getenv("USERNAME", "")
    return u.strip() or None

def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def _bootstrap_staff_user():
    username = _get_username()
    stored_hash = _get_stored_hash()
    if not username or stored_hash is None:
        return
    staff_col = get_staff_col()
    staff_col.create_index("username_norm", unique=True)
    if staff_col.count_documents({"username_norm": username.lower()}, limit=1) == 0:
        staff_col.insert_one({
            "username": username,
            "username_norm": username.lower(),
            "full_name": username.title(),
            "role": "admin",
            "password_hash": stored_hash.decode() if isinstance(stored_hash, bytes) else stored_hash,
            "active": True,
            "created_at": str(datetime.now()),
            "created_by": "bootstrap",
        })

def _find_staff_user(username: str) -> dict | None:
    if not username:
        return None
    try:
        _bootstrap_staff_user()
    except Exception:
        return None
    return get_staff_col().find_one({"username_norm": username.strip().lower(), "active": True})

def _requires_2fa() -> bool:
    env_flag = str(os.getenv("REQUIRE_2FA", "")).strip().lower() in ("1", "true", "yes")
    try:
        security = get_collection("settings").find_one({"_id": "security"}) or {}
        return env_flag or bool(security.get("require_2fa", False))
    except Exception:
        return env_flag

def _verify_2fa(code: str) -> bool:
    try:
        expected = st.secrets.get("ADMIN_2FA_CODE") or os.getenv("ADMIN_2FA_CODE", "")
    except Exception:
        expected = os.getenv("ADMIN_2FA_CODE", "")
    expected = str(expected).strip()
    return bool(expected) and hmac.compare_digest(str(code).strip(), expected)


def _check_lockout() -> tuple[bool, int]:
    """Returns (is_locked, seconds_remaining)."""
    attempts  = st.session_state.get("login_attempts", 0)
    lock_time = st.session_state.get("login_lock_until", 0)
    now       = time.time()
    if lock_time and now < lock_time:
        return True, int(lock_time - now)
    if lock_time and now >= lock_time:
        # Reset after lockout expires
        st.session_state.login_attempts   = 0
        st.session_state.login_lock_until = 0
    return False, 0


def _record_failure():
    attempts = st.session_state.get("login_attempts", 0) + 1
    st.session_state.login_attempts = attempts
    if attempts >= _MAX_ATTEMPTS:
        st.session_state.login_lock_until = time.time() + _LOCKOUT_SECS


def _authenticate_credentials(username: str, password: str) -> dict | None:
    staff = _find_staff_user(username)
    if staff:
        try:
            stored_hash = staff.get("password_hash", "").encode()
            if bcrypt.checkpw(password.encode(), stored_hash):
                return staff
        except Exception:
            return None

    stored_user = _get_username()
    stored_hash = _get_stored_hash()
    if stored_user is None or stored_hash is None:
        return None

    user_ok = hmac.compare_digest(username.encode(), stored_user.encode())
    try:
        pass_ok = bcrypt.checkpw(password.encode(), stored_hash)
    except Exception:
        pass_ok = False
    if user_ok and pass_ok:
        return {"username": stored_user, "full_name": stored_user.title(), "role": "admin", "active": True}
    return None

def _verify_credentials(username: str, password: str) -> bool:
    return _authenticate_credentials(username, password) is not None


# =====================================================
# ADMIN LOGIN
# =====================================================

def render_admin_login_strip():
    st.markdown("<div class='admin-strip'>", unsafe_allow_html=True)
    st.markdown("<div class='admin-strip-label'>◆ Admin Access</div>", unsafe_allow_html=True)

    # Credential config check (fail closed if neither staff users nor env admin exist)
    try:
        staff_configured = get_staff_col().count_documents({"active": True}, limit=1) > 0
    except Exception:
        staff_configured = False
    if not staff_configured and (_get_stored_hash() is None or _get_username() is None):
        with st.expander("Sign in to Admin Dashboard", expanded=False):
            st.error(
                "Admin credentials are not configured. "
                "Set USERNAME and PASSWORD (or PASSWORD_HASH) in st.secrets or environment variables."
            )
        st.markdown("</div>", unsafe_allow_html=True)
        return

    with st.expander("Sign in to Admin Dashboard", expanded=False):
        locked, secs_left = _check_lockout()
        if locked:
            st.error(f"Too many failed attempts. Try again in {secs_left // 60}m {secs_left % 60}s.")
            st.markdown("</div>", unsafe_allow_html=True)
            return

        attempts_left = _MAX_ATTEMPTS - st.session_state.get("login_attempts", 0)
        if attempts_left < _MAX_ATTEMPTS:
            st.warning(f"{attempts_left} attempt(s) remaining before lockout.")

        with st.form("admin_login_form"):
            u = st.text_input("Username", placeholder="username",   key="admin_u")
            p = st.text_input("Password", type="password",
                              placeholder="••••••••",                key="admin_p")
            twofa_required = _requires_2fa()
            twofa = st.text_input("2FA Code", type="password", placeholder="Required when enabled", key="admin_2fa") if twofa_required else ""
            submitted = st.form_submit_button("Sign In", use_container_width=True)

        if submitted:
            staff = _authenticate_credentials(u, p)
            if staff and (not twofa_required or _verify_2fa(twofa)):
                st.session_state.logged_in      = True
                st.session_state.username        = staff.get("username", u)
                st.session_state.role            = staff.get("role", "admin")
                st.session_state.staff_id        = str(staff.get("_id", "env-admin"))
                st.session_state.last_activity   = time.time()
                st.session_state.login_attempts  = 0
                st.session_state.login_lock_until = 0
                try:
                    get_staff_col().update_one(
                        {"username_norm": str(staff.get("username", u)).lower()},
                        {"$set": {"last_login_at": str(datetime.now())}},
                    )
                except Exception:
                    pass
                st.rerun()
            elif staff and twofa_required:
                _record_failure()
                st.error("Invalid or missing 2FA code.")
            else:
                _record_failure()
                locked2, secs_left2 = _check_lockout()
                if locked2:
                    st.error(f"Account locked for {secs_left2 // 60}m {secs_left2 % 60}s.")
                else:
                    st.error("Invalid credentials.")

    st.markdown("</div>", unsafe_allow_html=True)

# =====================================================
# ADMIN SIDEBAR
# =====================================================

def render_global_search():
    query = st.text_input("Global Search", placeholder="Customer, phone, sale ID, SKU, notes", key="global_search")
    if not query:
        return
    df = fetch_all()
    if df.empty:
        st.caption("No records to search.")
        return
    q = query.strip()
    mask = (
        df["customer_name"].astype(str).str.contains(q, case=False, na=False)
        | df["customer_phone"].astype(str).str.contains(q, case=False, na=False)
        | df["product_description"].astype(str).str.contains(q, case=False, na=False)
        | df["notes"].astype(str).str.contains(q, case=False, na=False)
        | df["sku"].astype(str).str.contains(q, case=False, na=False)
        | df["id"].astype(str).eq(q)
    )
    hits = df[mask].sort_values("sale_date", ascending=False).head(10)
    if hits.empty:
        st.caption("No matching customer or transaction.")
        return
    selected = st.selectbox(
        "Matches",
        hits["id"].tolist(),
        format_func=lambda x: f"#{x} — {hits[hits['id']==x]['customer_name'].iloc[0]} · ₹{hits[hits['id']==x]['total_amount'].iloc[0]:,.0f}",
        key="global_search_match",
    )
    if st.button("Open Match", key="global_search_open"):
        st.session_state.global_update_id = int(selected)
        st.session_state.nav_page = "Update Transaction"
        st.rerun()

def sidebar():
    with st.sidebar:
        # ── THEME TOGGLE ──────────────────────────────────────────────────
        theme_labels = {"system": "System Theme", "light": "Light Mode", "dark": "Dark Mode"}
        current_theme = st.session_state.get("theme", "system")
        selected_theme = st.selectbox(
            "Theme",
            ["system", "light", "dark"],
            index=["system", "light", "dark"].index(current_theme) if current_theme in ["system", "light", "dark"] else 0,
            format_func=lambda x: theme_labels[x],
            key="theme_select",
        )
        if selected_theme != current_theme:
            st.session_state.theme = selected_theme
            st.rerun()

        st.markdown("""
        <div class='sb-brand'>
            <div class='sb-logo' style='font-family:"DM Serif Display",serif'>Vinay</div>
            <div class='sb-mark'>Boutique Manager</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)

        df = fetch_all()
        m  = metrics(df)

        c1, c2 = st.columns(2)
        c1.metric("Pending", f"₹{m['pending']:,.0f}")
        c2.metric("Net",     f"₹{m['net_profit']:,.0f}")
        c1.metric("Sales",   m["sales"])
        c2.metric("Clients", m["customers"])

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)
        render_global_search()

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)

        pages = allowed_pages_for_current_role()
        if st.session_state.get("nav_page") not in pages:
            st.session_state.nav_page = pages[0]
        nav = st.radio("Navigation", pages, label_visibility="collapsed", key="nav_page")

        st.markdown("<div class='sb-sep'></div>", unsafe_allow_html=True)
        st.markdown(
            f"<div class='sb-user'>◆ {st.session_state.get('username','Admin').title()} · {current_role().title()}</div>",
            unsafe_allow_html=True,
        )
    return nav

# =====================================================
# ADMIN PAGES
# =====================================================

def page_dashboard():
    page_header("Dashboard", "Business Overview")
    all_df = fetch_all()
    m  = metrics(all_df)
    settings_col = get_db()["settings"]
    target_doc = settings_col.find_one({"_id": "monthly_target"}) or {}
    current_target = float(target_doc.get("amount", 0) or 0)

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Sales",      m["sales"])
    c2.metric("Billed",     f"₹{m['revenue']:,.0f}")
    c3.metric("Collected",  f"₹{m['collected']:,.0f}")
    c4.metric("Pending",    f"₹{m['pending']:,.0f}")
    c5.metric("Expenses",   f"₹{m['expenses']:,.0f}")
    c6.metric("Net Profit", f"₹{m['net_profit']:,.0f}")
    st.caption("Net Profit = merchandise profit minus refunds and recorded expenses such as rent, salaries, packaging, marketing, and vendor payments.")

    current_month = pd.Timestamp(date.today()).to_period("M").strftime("%Y-%m")
    month_sales = accounted_sales(all_df)
    month_revenue = 0
    if not month_sales.empty:
        month_revenue = month_sales[month_sales["sale_date"].dt.to_period("M").astype(str) == current_month]["total_amount"].sum()
    with st.expander("Monthly Target", expanded=current_target > 0):
        tc1, tc2 = st.columns([1, 2])
        with tc1:
            new_target = st.number_input("Target (₹)", min_value=0.0, value=current_target, step=5000.0, format="%.2f")
            if st.button("Save Target"):
                settings_col.update_one({"_id": "monthly_target"}, {"$set": {"amount": round(float(new_target), 2), "updated_at": str(datetime.now())}}, upsert=True)
                st.success("Monthly target saved.")
                st.rerun()
        with tc2:
            progress = min(month_revenue / current_target, 1.0) if current_target > 0 else 0.0
            st.metric("This Month", f"₹{month_revenue:,.0f}", f"{progress * 100:.1f}% of target" if current_target > 0 else "No target set")
            st.progress(progress)

    rule()

    if all_df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No sales yet.</div></div>", unsafe_allow_html=True)
        if st.button("Add your first sale", use_container_width=True):
            st.session_state.nav_page = "Add Sale"
            st.rerun()
        return

    df = accounted_sales(all_df)
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No recorded sales yet. Pending payments are still included in the Pending total.</div></div>", unsafe_allow_html=True)
        if st.button("Add your first sale", use_container_width=True):
            st.session_state.nav_page = "Add Sale"
            st.rerun()
        return

    df["month"] = df["sale_date"].dt.to_period("M").astype(str)

    with st.spinner("Loading dashboard charts..."):
        cl, cr = st.columns([3, 2])
        with cl:
            monthly = df.groupby("month").agg(revenue=("total_amount","sum"), profit=("profit","sum"), sales=("id","count")).reset_index()
            fig = go.Figure()
            fig.add_trace(go.Bar(x=monthly["month"], y=monthly["revenue"], name="Revenue", marker_color="rgba(46,111,216,0.4)", marker_line_color="#2E6FD8", marker_line_width=1))
            fig.add_trace(go.Scatter(x=monthly["month"], y=monthly["profit"], name="Profit", mode="lines+markers", line=dict(color="#7ADFA0", width=2), marker=dict(size=5, color="#7ADFA0")))
            styled_fig(fig, 300).update_layout(title="Monthly Revenue & Profit", barmode="overlay", legend=dict(orientation="h", y=1.18, x=0))
            st.plotly_chart(fig, use_container_width=True)

        with cr:
            paid    = (df["payment_received"] == 1).sum()
            pending = (df["payment_received"] == 0).sum()
            fig2 = go.Figure(go.Pie(labels=["Collected","Pending"], values=[paid, pending], hole=0.72, marker=dict(colors=["#2E6FD8","#0F1A2E"]), textfont=dict(size=11), hovertemplate="%{label}: %{value}<extra></extra>"))
            fig2.add_annotation(text=f"<b>{paid+pending}</b>", x=0.5, y=0.52, showarrow=False, font=dict(color="#E8EEF8", family="Playfair Display", size=28))
            fig2.add_annotation(text="sales", x=0.5, y=0.38, showarrow=False, font=dict(color="#3D5478", family="Jost", size=11))
            styled_fig(fig2, 300).update_layout(title="Payment Status", showlegend=True, legend=dict(orientation="h", y=-0.05, x=0.25))
            st.plotly_chart(fig2, use_container_width=True)

        cl2, cr2 = st.columns(2)
        with cl2:
            cat_rev = df.groupby("product_category")["total_amount"].sum().reset_index()
            fig3 = px.pie(cat_rev, values="total_amount", names="product_category", title="Revenue by Category", hole=0.55, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8","#9B9070","#A8C4F0"])
            styled_fig(fig3, 270); st.plotly_chart(fig3, use_container_width=True)

        with cr2:
            daily = df.set_index("sale_date")["total_amount"].resample("D").sum().reset_index()
            daily.columns = ["date","revenue"]
            daily["rolling"] = daily["revenue"].rolling(7, min_periods=1).mean()
            fig4 = go.Figure()
            fig4.add_trace(go.Bar(x=daily["date"], y=daily["revenue"], name="Daily", marker_color="rgba(46,111,216,0.25)", marker_line_width=0))
            fig4.add_trace(go.Scatter(x=daily["date"], y=daily["rolling"], name="7-day avg", line=dict(color="#2E6FD8", width=1.8)))
            styled_fig(fig4, 270).update_layout(title="Daily Revenue", legend=dict(orientation="h", y=1.18, x=0))
            st.plotly_chart(fig4, use_container_width=True)

    sec("Recent Transactions")
    dash_search = st.text_input("Quick customer lookup", placeholder="Search customer, phone, notes, or description", key="dashboard_quick_lookup")
    recent_source = df.copy()
    if dash_search:
        recent_source = recent_source[
            recent_source["customer_name"].str.contains(dash_search, case=False, na=False)
            | recent_source["customer_phone"].astype(str).str.contains(dash_search, case=False, na=False)
            | recent_source["notes"].astype(str).str.contains(dash_search, case=False, na=False)
            | recent_source["product_description"].astype(str).str.contains(dash_search, case=False, na=False)
        ]
    recent = recent_source.sort_values("sale_date", ascending=False).copy()
    recent = paginated_slice(recent, "dashboard_recent", default_size=10).copy()
    recent["sale_date"] = recent["sale_date"].dt.strftime("%d %b %Y")
    recent["Status"]    = recent["payment_received"].map({1:"Paid", 0:"Pending"})
    recent["Delayed"]   = recent["delay_status"].map({0:"—", 1:"Yes"})
    show = recent[["id","customer_name","sale_date","product_category","total_amount","profit","pending_amount","Status","Delayed"]].copy()
    show.columns = ["ID","Customer","Date","Category","Amount ₹","Profit ₹","Pending ₹","Status","Delayed"]
    st.dataframe(show, use_container_width=True, hide_index=True)

    rule()
    da, db, _ = st.columns([1, 1, 2])
    with da:
        st.download_button("Export CSV", data=df.assign(sale_date=df["sale_date"].astype(str)).to_csv(index=False), file_name=f"boutique_{date.today()}.csv", mime="text/csv", use_container_width=True)
    with db:
        st.download_button("Export Excel", data=to_excel(df), file_name=f"boutique_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)


def page_review():
    page_header("Accounts", "All Transactions")
    df = fetch_all()
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No transactions yet.</div></div>", unsafe_allow_html=True)
        return

    with st.expander("Filter & Sort", expanded=True):
        c1, c2, c3, c4, c0 = st.columns(5)
        with c1: srch  = st.text_input("Customer / Phone / Notes", key="review_filter_search")
        with c2: catf  = st.selectbox("Category",   ["All"] + CATEGORIES, key="review_filter_category")
        with c3: payf  = st.selectbox("Payment",    ["All","Paid","Pending"], key="review_filter_payment")
        with c4: dlayf = st.selectbox("Delay Flag", ["All","On Time","Delayed"], key="review_filter_delay")
        with c0: typef = st.selectbox("Type", ["All","Sale","Pending Payment","Return","Expense"], key="review_filter_type")
        c5, c6, c7, c8 = st.columns(4)
        with c5: sortby = st.selectbox("Sort By", ["Date ↓","Date ↑","Amount ↓","Pending ↓","Profit ↓"], key="review_filter_sort")
        with c6: preset = st.selectbox("Range", ["Last 90 days","Last 7 days","Last 30 days","This month","This year","Custom"], key="review_filter_range")
        today = date.today()
        preset_from = {
            "Last 7 days": today - timedelta(days=7),
            "Last 30 days": today - timedelta(days=30),
            "Last 90 days": today - timedelta(days=90),
            "This month": today.replace(day=1),
            "This year": today.replace(month=1, day=1),
        }.get(preset, today - timedelta(days=90))
        with c7: d_from = st.date_input("From", value=preset_from, disabled=(preset != "Custom"), key="review_filter_from")
        with c8: d_to   = st.date_input("To",   value=today, disabled=(preset != "Custom"), key="review_filter_to")

    fdf = df.copy()
    if srch:
        mask = (
            fdf["customer_name"].str.contains(srch, case=False, na=False)
            | fdf["customer_phone"].astype(str).str.contains(srch, case=False, na=False)
            | fdf["notes"].astype(str).str.contains(srch, case=False, na=False)
            | fdf["product_description"].astype(str).str.contains(srch, case=False, na=False)
        )
        fdf = fdf[mask]
    if catf  != "All": fdf = fdf[fdf["product_category"] == catf]
    if typef != "All":
        fdf = fdf[fdf["transaction_type"] == typef.lower().replace(" ", "_")]
    if payf  == "Paid":     fdf = fdf[fdf["payment_received"] == 1]
    elif payf == "Pending": fdf = fdf[fdf["payment_received"] == 0]
    if dlayf == "On Time":  fdf = fdf[fdf["delay_status"] == 0]
    elif dlayf == "Delayed": fdf = fdf[fdf["delay_status"] == 1]
    fdf = fdf[(fdf["sale_date"] >= pd.Timestamp(d_from)) & (fdf["sale_date"] <= pd.Timestamp(d_to))]
    sm = {"Date ↓":("sale_date",False),"Date ↑":("sale_date",True),"Amount ↓":("total_amount",False),"Pending ↓":("pending_amount",False),"Profit ↓":("profit",False)}
    sc, sa = sm[sortby]
    fdf = fdf.sort_values(sc, ascending=sa)

    rule_sm()
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Transactions", len(fdf))
    m2.metric("Revenue",      f"₹{fdf['total_amount'].sum():,.0f}")
    m3.metric("Profit",       f"₹{fdf['profit'].sum():,.0f}")
    m4.metric("Pending",      f"₹{fdf['pending_amount'].sum():,.0f}")
    m5.metric("Avg Margin",   f"{fdf['margin'].mean():.1f}%" if not fdf.empty else "—")
    rule_sm()

    sec("Bulk Actions")
    if fdf.empty:
        st.info("No transactions match the current filters.")
    else:
        bulk_ids = st.multiselect(
            "Select Transactions",
            fdf["id"].tolist(),
            format_func=lambda x: f"#{x} — {fdf[fdf['id']==x]['customer_name'].iloc[0]}",
            key="review_bulk_ids",
        )
        if bulk_ids:
            ba, bb, bc, bd = st.columns(4)
            with ba:
                bulk_method = st.selectbox("Payment Method", PAYMENT_METHODS, key="review_bulk_payment_method")
            with bb:
                if st.button("Mark Paid", key="review_bulk_paid"):
                    updated = 0
                    for rid in bulk_ids:
                        ok, _ = mark_transaction_paid(rid, bulk_method, "Bulk mark paid from review")
                        updated += int(ok)
                    invalidate_cache()
                    st.success(f"{updated} transaction(s) marked paid.")
                    st.rerun()
            with bc:
                if st.button("Flag Delayed", key="review_bulk_flag"):
                    updated = 0
                    for rid in bulk_ids:
                        ok, _ = update_record_with_audit(rid, set_fields={"delay_status": 1}, action="bulk_flag_delayed")
                        updated += int(ok)
                    invalidate_cache()
                    st.success(f"{updated} transaction(s) flagged.")
                    st.rerun()
            with bd:
                confirm_bulk_delete = st.checkbox("Confirm trash", key="review_bulk_delete_confirm")
                if st.button("Move to Trash", key="review_bulk_delete"):
                    if not confirm_bulk_delete:
                        st.error("Tick confirm trash before moving records.")
                    else:
                        deleted = 0
                        for rid in bulk_ids:
                            ok, _ = soft_delete_transaction(rid, "Bulk action from review")
                            deleted += int(ok)
                        invalidate_cache()
                        st.success(f"{deleted} transaction(s) moved to trash.")
                        st.rerun()

    show = fdf[["id","transaction_type","customer_name","customer_phone","sale_date","product_category","buying_price","total_amount","profit","amount_paid","pending_amount","payment_method","delay_status","payment_received"]].copy()
    show["sale_date"]        = show["sale_date"].dt.strftime("%d %b %Y")
    show["delay_status"]     = show["delay_status"].map({0:"—", 1:"Yes"})
    show["payment_received"] = show["payment_received"].map({0:"Pending", 1:"Paid"})
    show.columns = ["ID","Type","Customer","Phone","Date","Category","Buy ₹","Amount ₹","Profit ₹","Paid ₹","Pending ₹","Method","Delayed","Status"]
    st.dataframe(paginated_slice(show, "review_table"), use_container_width=True, hide_index=True)

    dc, de, _ = st.columns([1,1,2])
    with dc:
        st.download_button("Export CSV", data=fdf.assign(sale_date=fdf["sale_date"].astype(str)).to_csv(index=False), file_name=f"accounts_{date.today()}.csv", mime="text/csv", use_container_width=True)
    with de:
        st.download_button("Export Excel", data=to_excel(fdf), file_name=f"accounts_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    sec("Mark Payments")
    pend = fdf[fdf["pending_amount"] > 0].sort_values("pending_amount", ascending=False)
    if pend.empty:
        st.success("All payments received for current filter.")
    else:
        st.markdown(f"<span class='badge badge-gold'>{len(pend)} pending — ₹{pend['pending_amount'].sum():,.0f}</span>", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        pend_view = paginated_slice(pend, "review_pending", default_size=10)
        for _, row in pend_view.iterrows():
            ca, cb, cc, cd, cm, ce = st.columns([3,2,1.5,1.5,1.6,1])
            ca.write(f"**{row['customer_name']}** · {row['product_category']}")
            cb.write(f"₹{row['pending_amount']:,.2f} pending")
            cc.write(row["sale_date"].strftime("%d %b %Y") if pd.notna(row["sale_date"]) else "—")
            with cd:
                pay_now = st.number_input(
                    "Pay",
                    min_value=0.0,
                    max_value=float(row["pending_amount"]),
                    value=float(row["pending_amount"]),
                    step=100.0,
                    format="%.2f",
                    key=f"pay_{row['id']}",
                    label_visibility="collapsed",
                )
            with cm:
                pay_method = st.selectbox(
                    "Method",
                    PAYMENT_METHODS,
                    index=PAYMENT_METHODS.index(row.get("payment_method")) if row.get("payment_method") in PAYMENT_METHODS else 0,
                    key=f"pay_method_{row['id']}",
                    label_visibility="collapsed",
                )
            with ce:
                if st.button("Add", key=f"p_{row['id']}"):
                    if pay_now <= 0:
                        st.warning("Enter a payment amount.")
                    else:
                        ok, status = apply_payment_update(row["id"], pay_now, pay_method, "Payment update", expected_version=safe_int(row.get("version"), 1))
                        if not ok and status == "conflict":
                            st.warning("This record changed while you were editing. Refresh and try again.")
                        elif not ok:
                            st.warning("Payment was not applied.")
                        else:
                            invalidate_cache(); st.rerun()
                if st.button("Flag", key=f"f_{row['id']}"):
                    ok, status = update_record_with_audit(row["id"], set_fields={"delay_status": 1}, action="flag_delayed", expected_version=safe_int(row.get("version"), 1))
                    if not ok and status == "conflict":
                        st.warning("This record changed while you were editing. Refresh and try again.")
                    else:
                        invalidate_cache(); st.rerun()


def page_update():
    page_header("Update", "Edit or Delete a Record")
    c1, c2 = st.columns([2,1])
    with c1: sname = st.text_input("Search by Customer Name", key="update_search_name")
    default_sid = int(st.session_state.get("global_update_id", 0) or 0)
    with c2: sid   = st.number_input("Or by Sale ID", min_value=0, step=1, value=default_sid, key="update_search_id")

    if not sname and sid == 0:
        st.info("Enter a customer name or sale ID to search.")
        return

    q = active_filter({"customer_name": {"$regex": sname, "$options":"i"}} if sname else {"id": int(sid)})
    docs = list(get_col().find(q, {"_id":0}))
    if not docs:
        st.warning("No matching transaction found.")
        return

    df = pd.DataFrame(docs)
    for c in ["buying_price","selling_price","amount_paid","pending_amount"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    preview_cols = [c for c in ["id","customer_name","sale_date","product_category","selling_price","pending_amount","payment_received"] if c in df.columns]
    preview = df[preview_cols].copy()
    if "payment_received" in preview.columns:
        preview["payment_received"] = preview["payment_received"].map({0:"Pending",1:"Paid"})
    st.dataframe(preview, use_container_width=True, hide_index=True)

    sel = st.selectbox("Select ID to Edit", df["id"].tolist(), format_func=lambda x: f"#{x} — {df[df['id']==x]['customer_name'].values[0]}")
    row = df[df["id"] == sel].iloc[0]
    rule_sm()

    with st.form("update_form"):
        sec("Customer & Product")
        c1, c2, c3 = st.columns(3)
        with c1:
            nn = st.text_input("Customer Name", value=str(row.get("customer_name","")))
            n_phone = st.text_input("Phone",     value=str(row.get("customer_phone","")))
        with c2:
            ci  = CATEGORIES.index(row["product_category"]) if row.get("product_category") in CATEGORIES else 0
            nc  = st.selectbox("Category", CATEGORIES, index=ci)
            nv  = st.text_input("Vendor", value=str(row.get("vendor","")))
        with c3:
            try:    existing_date = pd.to_datetime(row.get("sale_date")).date()
            except: existing_date = date.today()
            new_date = st.date_input("Sale Date", value=existing_date)
            nqty = st.number_input("Quantity", min_value=1, step=1, value=int(row.get("quantity",1)))

        ndesc = st.text_area("Description", value=str(row.get("product_description","")), height=60)
        sec("Pricing & Payment")
        pr1, pr2, pr3, pr4 = st.columns(4)
        with pr1: nb  = st.number_input("Buying Price (₹)",  value=float(row["buying_price"]),  min_value=0.0, step=100.0, format="%.2f")
        with pr2: ns  = st.number_input("Selling Price (₹)", value=float(row["selling_price"]), min_value=0.0, step=100.0, format="%.2f")
        with pr3: npa = st.number_input("Amount Paid (₹)",   value=float(row["amount_paid"]),   min_value=0.0, step=100.0, format="%.2f")
        with pr4:
            pi  = PAYMENT_METHODS.index(row["payment_method"]) if row.get("payment_method") in PAYMENT_METHODS else 0
            npm = st.selectbox("Payment Method", PAYMENT_METHODS, index=pi)

        nd     = st.checkbox("Mark as Delayed", value=bool(row.get("delay_status",0)))
        nnotes = st.text_area("Notes", value=str(row.get("notes","")), height=60)

        ntotal  = round(ns * nqty, 2)
        npend   = max(round(ntotal - npa, 2), 0.0)
        nprofit = round((ns - nb) * nqty, 2)
        m1, m2, m3 = st.columns(3)
        m1.metric("Updated Pending", f"₹{npend:,.2f}")
        m2.metric("Updated Profit",  f"₹{nprofit:,.2f}")
        m3.metric("Updated Margin",  f"{(nprofit/ntotal*100 if ntotal>0 else 0):.1f}%")

        confirm_delete = st.checkbox("Confirm delete for this transaction")
        bu, bd = st.columns(2)
        with bu: upd = st.form_submit_button("Save Changes",       use_container_width=True)
        with bd: dlt = st.form_submit_button("Delete Transaction",  use_container_width=True)

        if upd:
            errs = []
            if not nn.strip():
                errs.append("Customer name is required.")
            if n_phone.strip() and not is_valid_indian_phone(n_phone):
                errs.append("Enter a valid Indian phone number.")
            if npa > ntotal:
                errs.append("Amount paid cannot exceed total sale value.")
            row_type = row.get("transaction_type", "sale") or "sale"
            if errs:
                for e in errs: st.error(e)
            else:
                old_paid = float(row.get("amount_paid", 0) or 0)
                old_qty = int(row.get("quantity", 1) or 1)
                old_cat = str(row.get("product_category", ""))
                old_desc = str(row.get("product_description", ""))
                set_fields = {
                    "customer_name": nn.strip(), "customer_phone": n_phone.strip(),
                    "sale_date": str(new_date), "product_category": nc,
                    "vendor": nv.strip(), "product_description": ndesc.strip(),
                    "quantity": nqty, "buying_price": round(nb,2),
                    "selling_price": round(ns,2), "amount_paid": round(npa,2),
                    "pending_amount": npend, "delay_status": int(nd),
                    "payment_method": npm, "notes": nnotes.strip(),
                    "payment_received": 1 if npend==0 else 0,
                }
                push_fields = None
                if npa > old_paid:
                    push_fields = {"payment_history": {"amount": round(npa - old_paid, 2), "method": npm, "note": "Payment adjusted during update", "paid_at": str(datetime.now()), "recorded_by": current_user()}}
                elif npa < old_paid:
                    push_fields = {"payment_history": {"amount": round(old_paid - npa, 2), "method": npm, "note": "Payment correction reduced paid amount", "paid_at": str(datetime.now()), "recorded_by": current_user()}}
                ok, status = update_record_with_audit(sel, set_fields=set_fields, push_fields=push_fields, action="transaction_update", expected_version=safe_int(row.get("version"), 1))
                if not ok and status == "conflict":
                    st.warning("This transaction changed after you loaded it. Refresh and review the latest values before saving.")
                elif not ok:
                    st.warning("Transaction could not be updated.")
                else:
                    if row_type == "sale":
                        if old_cat != nc or old_desc.strip().lower() != ndesc.strip().lower():
                            increment_inventory(old_cat, old_qty, old_desc, str(row.get("vendor", "")), float(row.get("buying_price", 0) or 0), str(row.get("sku", "")))
                            if not decrement_inventory_for_sale(nc, nqty, ndesc.strip(), str(row.get("sku", ""))):
                                st.warning("Transaction updated, but the new inventory item did not have enough stock to deduct.")
                        elif nqty > old_qty:
                            if not decrement_inventory_for_sale(nc, nqty - old_qty, ndesc.strip(), str(row.get("sku", ""))):
                                st.warning("Transaction updated, but extra quantity was not available in inventory.")
                        elif nqty < old_qty:
                            increment_inventory(nc, old_qty - nqty, ndesc.strip(), nv.strip(), nb, str(row.get("sku", "")))
                    invalidate_cache(); st.success("Transaction updated."); st.rerun()
        if dlt:
            if not confirm_delete:
                st.error("Tick the confirmation checkbox before deleting.")
            else:
                ok, status = soft_delete_transaction(sel, "Manual trash from update page", expected_version=safe_int(row.get("version"), 1))
                if not ok and status == "conflict":
                    st.warning("This transaction changed after you loaded it. Refresh and try again.")
                elif not ok:
                    st.warning("Transaction could not be moved to trash.")
                else:
                    invalidate_cache(); st.success("Transaction moved to trash. You can recover it from Staff & Audit."); st.rerun()


def page_customers():
    page_header("Customers", "All Clients")
    df = fetch_all()
    if df.empty:
        st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No customers yet.</div></div>", unsafe_allow_html=True)
        if st.button("Add your first sale", use_container_width=True):
            st.session_state.nav_page = "Add Sale"
            st.rerun()
        return

    sales_df = accounted_sales(df)
    pending_by_customer = df.groupby("customer_name")["pending_amount"].sum()
    if sales_df.empty:
        st.info("No sale customers yet. Pending credits will appear once linked to sales or collected.")
        return

    summ = (sales_df.groupby("customer_name").agg(
        phone=("customer_phone","first"), transactions=("id","count"),
        spent=("total_amount","sum"), pending=("pending_amount","sum"),
        last_visit=("sale_date","max"), profit=("profit","sum"),
    ).reset_index())
    summ["pending"] = summ["customer_name"].map(pending_by_customer).fillna(0)
    summ["last_visit"] = pd.to_datetime(summ["last_visit"]).dt.strftime("%d %b %Y")
    summ = summ.sort_values("spent", ascending=False).reset_index(drop=True)
    summ["tier"] = summ["spent"].apply(lambda v: loyalty_for_spend(float(v))["tier"])
    summ["benefit"] = summ["spent"].apply(lambda v: f"{loyalty_for_spend(float(v))['discount_pct']:.0f}% discount")
    summ["points_rate"] = summ["spent"].apply(lambda v: loyalty_for_spend(float(v))["points_rate"])

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Customers", len(summ))
    m2.metric("Avg Spend",       f"₹{summ['spent'].mean():,.0f}")
    m3.metric("With Pending",    len(summ[summ["pending"] > 0]))
    m4.metric("Total Revenue",   f"₹{summ['spent'].sum():,.0f}")

    rule_sm()
    c1, c2 = st.columns([2,1])
    with c1: srch   = st.text_input("Search Customer", key="customers_search")
    with c2: tier_f = st.selectbox("Tier", ["All","Bronze","Silver","Gold","Platinum"], key="customers_tier")

    view = summ.copy()
    if srch:     view = view[view["customer_name"].str.contains(srch, case=False, na=False)]
    if tier_f != "All": view = view[view["tier"] == tier_f]

    disp = view.rename(columns={"customer_name":"Customer","phone":"Phone","transactions":"Visits","spent":"Total Spent ₹","pending":"Pending ₹","last_visit":"Last Visit","profit":"Profit ₹","tier":"Tier","benefit":"Benefit","points_rate":"Points / ₹100"})
    disp_page = paginated_slice(disp, "customers_table")
    st.dataframe(disp_page.style.format({"Total Spent ₹":"₹{:,.0f}","Pending ₹":"₹{:,.0f}","Profit ₹":"₹{:,.0f}"}), use_container_width=True, hide_index=True)

    sec("Update Customer Phone")
    with st.form("customer_phone_form"):
        pc1, pc2 = st.columns(2)
        with pc1:
            phone_customer = st.selectbox("Customer", summ["customer_name"].tolist(), key="phone_update_customer")
        with pc2:
            current_phone = str(summ[summ["customer_name"] == phone_customer]["phone"].iloc[0] or "") if phone_customer else ""
            new_phone = st.text_input("Phone", value=current_phone, placeholder="+91 XXXXXXXXXX")
        if st.form_submit_button("Update Phone Across Records", use_container_width=True):
            if new_phone.strip() and not is_valid_indian_phone(new_phone):
                st.error("Enter a valid Indian phone number.")
            else:
                docs = list(get_col().find(active_filter({"customer_name": {"$regex": f"^{re.escape(phone_customer)}$", "$options": "i"}}), {"id": 1, "version": 1}))
                updated = 0
                for doc in docs:
                    ok, _ = update_record_with_audit(
                        doc["id"],
                        set_fields={"customer_phone": new_phone.strip()[:20]},
                        action="customer_phone_update",
                        expected_version=safe_int(doc.get("version"), 1),
                    )
                    updated += int(ok)
                invalidate_cache()
                st.success(f"Phone updated on {updated} record(s).")
                st.rerun()

    dc, de = st.columns(2)
    with dc:
        st.download_button("Export CSV", data=disp.to_csv(index=False), file_name=f"customers_{date.today()}.csv", mime="text/csv", use_container_width=True)
    with de:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w: disp.to_excel(w, index=False)
        out.seek(0)
        st.download_button("Export Excel", data=out, file_name=f"customers_{date.today()}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    sec("Purchase History")
    chosen = st.selectbox("Select Customer", summ["customer_name"].tolist())
    if chosen:
        hist = df[df["customer_name"] == chosen].sort_values("sale_date", ascending=False).copy()
        hist["status"]   = hist["payment_received"].map({0:"Pending",1:"Paid"})
        hist["sale_date"] = hist["sale_date"].dt.strftime("%d %b %Y")
        h1, h2, h3, h4 = st.columns(4)
        h1.metric("Visits",      len(hist))
        h2.metric("Total Spent", f"₹{accounted_sales(hist)['total_amount'].sum():,.0f}")
        h3.metric("Pending",     f"₹{hist['pending_amount'].sum():,.0f}")
        h4.metric("Profit",      f"₹{hist['profit'].sum():,.0f}")
        cols = [c for c in ["sale_date","product_category","product_description","total_amount","amount_paid","pending_amount","payment_method","status"] if c in hist.columns]
        show = hist[cols].copy()
        show.columns = ["Date","Category","Description","Price ₹","Paid ₹","Pending ₹","Method","Status"][:len(cols)]
        st.dataframe(paginated_slice(show, "customer_history", default_size=10), use_container_width=True, hide_index=True)
        if len(hist) > 1:
            hs = accounted_sales(df[df["customer_name"]==chosen]).sort_values("sale_date").copy()
            hs["cumulative"] = hs["total_amount"].cumsum()
            fig = px.line(hs, x="sale_date", y="cumulative", title=f"Cumulative Spend — {chosen}", markers=True)
            fig.update_traces(line_color="#2E6FD8", marker_color="#4D8AE8", marker_size=5)
            styled_fig(fig, 230); st.plotly_chart(fig, use_container_width=True)


def page_analytics():
    page_header("Analytics", "Business Intelligence")
    df = accounted_sales(fetch_all())
    if df.empty:
        st.info("No recorded sales available.")
        return

    df["month"] = df["sale_date"].dt.to_period("M").astype(str)
    df["dow"]   = df["sale_date"].dt.day_name()

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Revenue",       f"₹{df['total_amount'].sum():,.0f}")
    k2.metric("Profit",        f"₹{df['profit'].sum():,.0f}")
    k3.metric("Avg Order",     f"₹{df['total_amount'].mean():,.0f}")
    k4.metric("Avg Margin",    f"{df['margin'].mean():.1f}%")
    k5.metric("Delayed Count", int((df["delay_status"]==1).sum()))

    rule()
    t1, t2, t3, t4, t5 = st.tabs(["Trends","Customers","Categories","Payments","Top Items"])

    with t1:
        c1, c2 = st.columns(2)
        with c1:
            monthly = df.groupby("month").agg(revenue=("total_amount","sum"), profit=("profit","sum")).reset_index()
            fig = go.Figure()
            fig.add_trace(go.Bar(x=monthly["month"], y=monthly["revenue"], name="Revenue", marker_color="rgba(46,111,216,0.4)", marker_line_color="#2E6FD8", marker_line_width=1))
            fig.add_trace(go.Scatter(x=monthly["month"], y=monthly["profit"], name="Profit", mode="lines+markers", line=dict(color="#7ADFA0", width=2), marker=dict(size=5)))
            styled_fig(fig).update_layout(title="Revenue & Profit by Month", barmode="overlay", legend=dict(orientation="h", y=1.18, x=0))
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            daily = df.set_index("sale_date")["total_amount"].resample("D").sum().reset_index()
            daily.columns = ["date","revenue"]
            fig2 = px.area(daily, x="date", y="revenue", title="Daily Revenue")
            fig2.update_traces(fillcolor="rgba(46,111,216,0.12)", line_color="#2E6FD8", line_width=1.5)
            styled_fig(fig2); st.plotly_chart(fig2, use_container_width=True)
        c3, c4 = st.columns(2)
        with c3:
            dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            dow = df.groupby("dow").agg(sales=("id","count"), revenue=("total_amount","sum")).reset_index()
            dow["dow"] = pd.Categorical(dow["dow"], categories=dow_order, ordered=True)
            dow = dow.sort_values("dow")
            fig3 = px.bar(dow, x="dow", y="sales", title="Sales by Day of Week", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
            styled_fig(fig3); st.plotly_chart(fig3, use_container_width=True)
        with c4:
            monthly["MoM Growth %"] = monthly["revenue"].pct_change()*100
            fig4 = px.bar(monthly.dropna(), x="month", y="MoM Growth %", title="Month-over-Month Growth", color="MoM Growth %", color_continuous_scale=[[0,"#C05060"],[0.5,"#0F1A2E"],[1,"#7ADFA0"]])
            styled_fig(fig4); st.plotly_chart(fig4, use_container_width=True)

    with t2:
        c1, c2 = st.columns(2)
        with c1:
            top_c = df.groupby("customer_name")["total_amount"].sum().nlargest(10).reset_index()
            fig5 = px.bar(top_c, x="total_amount", y="customer_name", orientation="h", title="Top 10 Customers by Revenue", color="total_amount", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
            styled_fig(fig5); fig5.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig5, use_container_width=True)
        with c2:
            cp = df.groupby("customer_name")["pending_amount"].sum()
            cp = cp[cp > 0].nlargest(10).reset_index()
            if not cp.empty:
                fig6 = px.bar(cp, x="pending_amount", y="customer_name", orientation="h", title="Top Customers by Pending", color="pending_amount", color_continuous_scale=[[0,"#070C18"],[1,"#C05060"]])
                styled_fig(fig6); fig6.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig6, use_container_width=True)
            else:
                st.success("No pending amounts.")
        cust_stats = df.groupby("customer_name").agg(visits=("id","count"), revenue=("total_amount","sum"), avg_order=("total_amount","mean")).reset_index()
        fig_scatter = px.scatter(cust_stats, x="visits", y="revenue", size="avg_order", hover_name="customer_name", title="Customer Value Matrix", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
        styled_fig(fig_scatter, 330); st.plotly_chart(fig_scatter, use_container_width=True)
        seg = df.groupby("customer_name").agg(spend=("total_amount","sum")).reset_index()
        seg["tier"] = pd.cut(seg["spend"], bins=[0,5000,20000,50000,float("inf")], labels=["Bronze","Silver","Gold","Platinum"])
        sec("Customer Tier Distribution")
        sg = seg.groupby("tier", observed=True).agg(customers=("customer_name","count"), total=("spend","sum")).reset_index()
        sg.columns = ["Tier","Customers","Total Spend ₹"]
        st.dataframe(sg, use_container_width=True, hide_index=True)

    with t3:
        c1, c2 = st.columns(2)
        with c1:
            cd = df.groupby("product_category").size().reset_index(name="count")
            fig7 = px.pie(cd, values="count", names="product_category", title="Sales Volume by Category", hole=0.55, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#E08090","#1A3D80","#3D9A6C","#4A9AC8","#9B9070","#A8C4F0"])
            styled_fig(fig7); st.plotly_chart(fig7, use_container_width=True)
        with c2:
            cp2 = df.groupby("product_category").agg(profit=("profit","sum"), revenue=("total_amount","sum")).reset_index()
            cp2["margin"] = (cp2["profit"]/cp2["revenue"]*100).round(1)
            fig8 = px.bar(cp2, x="product_category", y="profit", title="Profit by Category", color="margin", color_continuous_scale=[[0,"#070C18"],[1,"#7ADFA0"]])
            styled_fig(fig8); st.plotly_chart(fig8, use_container_width=True)
        cm = df.groupby(["month","product_category"])["total_amount"].sum().unstack(fill_value=0)
        if not cm.empty:
            fig9 = px.imshow(cm.T, title="Category × Month Heatmap", color_continuous_scale=[[0,"#070C18"],[0.4,"#1A3D80"],[1,"#2E6FD8"]], aspect="auto")
            styled_fig(fig9, 300); st.plotly_chart(fig9, use_container_width=True)

    with t4:
        c1, c2 = st.columns(2)
        with c1:
            pm = df.groupby("payment_method").size().reset_index(name="count")
            fig10 = px.pie(pm, values="count", names="payment_method", title="Payment Method Distribution", hole=0.58, color_discrete_sequence=["#2E6FD8","#4D8AE8","#7ADFA0","#8BACD8","#1A3D80","#E08090"])
            styled_fig(fig10); st.plotly_chart(fig10, use_container_width=True)
        with c2:
            ps = df.groupby("payment_received").agg(count=("id","count"), total=("pending_amount","sum")).reset_index()
            ps["label"] = ps["payment_received"].map({0:"Pending",1:"Received"})
            fig11 = px.bar(ps, x="label", y="count", title="Payment Status", color="label", color_discrete_map={"Pending":"#2E6FD8","Received":"#7ADFA0"})
            styled_fig(fig11); st.plotly_chart(fig11, use_container_width=True)
        aged = df[df["pending_amount"] > 0].copy()
        if not aged.empty:
            today_ts = pd.Timestamp(date.today())
            aged["days"] = (today_ts - aged["sale_date"]).dt.days
            aged["bucket"] = pd.cut(aged["days"], bins=[0,7,15,30,60,9999], labels=["0–7d","8–15d","16–30d","31–60d","60d+"])
            ag = aged.groupby("bucket", observed=True)["pending_amount"].sum().reset_index()
            fig12 = px.bar(ag, x="bucket", y="pending_amount", title="Pending — Aging Buckets", color="pending_amount", color_continuous_scale=[[0,"#2E6FD8"],[1,"#C05060"]])
            styled_fig(fig12); st.plotly_chart(fig12, use_container_width=True)
        else:
            st.success("No pending payments.")

    with t5:
        c1, c2 = st.columns(2)
        with c1:
            if "vendor" in df.columns:
                vd = (df[df["vendor"].astype(str).str.strip() != ""].groupby("vendor").agg(revenue=("total_amount","sum"), items=("id","count")).nlargest(10,"revenue").reset_index())
                if not vd.empty:
                    fig13 = px.bar(vd, x="revenue", y="vendor", orientation="h", title="Top Vendors by Revenue", color="revenue", color_continuous_scale=[[0,"#070C18"],[1,"#2E6FD8"]])
                    styled_fig(fig13); fig13.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig13, use_container_width=True)
                else:
                    st.info("Add vendor names to see this chart.")
        with c2:
            if "product_description" in df.columns:
                pd2 = df[df["product_description"].astype(str).str.strip() != ""].copy()
                if not pd2.empty:
                    tm = (pd2.groupby("product_description").agg(margin=("margin","mean"), revenue=("total_amount","sum")).nlargest(10,"margin").reset_index())
                    tm["product_description"] = tm["product_description"].str[:30]
                    fig14 = px.bar(tm, x="margin", y="product_description", orientation="h", title="Top Products by Margin %", color="margin", color_continuous_scale=[[0,"#070C18"],[1,"#7ADFA0"]])
                    styled_fig(fig14); fig14.update_layout(yaxis=dict(autorange="reversed")); st.plotly_chart(fig14, use_container_width=True)
                else:
                    st.info("Add product descriptions to see this chart.")


def page_reminders():
    page_header("Reminders", "Payment Follow-ups")
    df = fetch_all()
    if df.empty:
        st.info("No data available.")
        return

    today_ts  = pd.Timestamp(date.today())
    df["days_old"] = (today_ts - df["sale_date"]).dt.days
    overdue_count = len(df[(df["pending_amount"] > 0) & (df["days_old"] > 30)])
    flagged_count = int((df["delay_status"] == 1).sum())

    if overdue_count or flagged_count:
        bc = st.columns(2)
        if overdue_count: bc[0].error(f"{overdue_count} payments overdue (30+ days)")
        if flagged_count: bc[1].warning(f"{flagged_count} transactions flagged")
    else:
        st.success("All clear — no overdue or flagged payments.")

    rule_sm()
    t1, t2, t3, t4 = st.tabs(["Overdue (30d+)","Flagged","High Value","Upcoming"])

    with t1:
        ov = df[(df["pending_amount"] > 0) & (df["days_old"] > 30)].sort_values("days_old", ascending=False)
        if ov.empty:
            st.success("No overdue payments.")
        else:
            st.warning(f"{len(ov)} overdue — ₹{ov['pending_amount'].sum():,.0f} total")
            st.caption("WhatsApp opens a message for staff to send. Record Follow-up only logs that a reminder was attempted.")
            bulk_overdue = st.multiselect(
                "Bulk overdue actions",
                ov["id"].tolist(),
                format_func=lambda x: f"#{x} — {ov[ov['id']==x]['customer_name'].iloc[0]}",
                key="reminders_bulk_overdue",
            )
            if bulk_overdue:
                bo1, bo2, bo3 = st.columns(3)
                with bo1:
                    bulk_method = st.selectbox("Payment Method", PAYMENT_METHODS, key="reminders_bulk_method")
                with bo2:
                    if st.button("Mark Selected Paid", key="reminders_bulk_paid"):
                        updated = 0
                        for rid in bulk_overdue:
                            ok, _ = mark_transaction_paid(rid, bulk_method, "Bulk mark paid from reminders")
                            updated += int(ok)
                        invalidate_cache()
                        st.success(f"{updated} transaction(s) marked paid.")
                        st.rerun()
                with bo3:
                    if st.button("Record Follow-ups", key="reminders_bulk_log"):
                        for rid in bulk_overdue:
                            log_reminder(rid, "WhatsApp")
                        invalidate_cache()
                        st.success(f"Follow-up logged for {len(bulk_overdue)} transaction(s).")
                        st.rerun()
            for _, r in ov.iterrows():
                with st.expander(f"{r['customer_name']}  ·  ₹{r['pending_amount']:,.0f}  ·  {int(r['days_old'])} days"):
                    ca, cb, cc, cd, ce = st.columns([2,2,1,1,1.5])
                    ca.write(r["sale_date"].strftime("%d %b %Y"))
                    cb.write(r.get("product_category","—"))
                    with cc:
                        if st.button("Mark Paid", key=f"op_{r['id']}"):
                            ok, status = mark_transaction_paid(r["id"], r.get("payment_method", ""), "Marked paid from reminders", expected_version=safe_int(r.get("version"), 1))
                            if not ok and status == "conflict":
                                st.warning("This record changed while you were viewing it. Refresh and try again.")
                            else:
                                invalidate_cache(); st.rerun()
                    with cd:
                        link = whatsapp_link(r.get("customer_phone", ""), r.get("customer_name", ""), float(r.get("pending_amount", 0) or 0))
                        if link:
                            st.link_button("WhatsApp", link, use_container_width=True)
                        else:
                            st.caption("No valid phone")
                    with ce:
                        if st.button("Record Follow-up", key=f"or_{r['id']}"):
                            log_reminder(r["id"])
                            st.toast(f"Follow-up logged for {r['customer_name']}. No message was sent automatically.")

    with t2:
        dl = df[df["delay_status"] == 1].sort_values("pending_amount", ascending=False)
        if dl.empty:
            st.success("No flagged payments.")
        else:
            st.error(f"{len(dl)} flagged — ₹{dl['pending_amount'].sum():,.0f}")
            show = dl[["customer_name","sale_date","product_category","total_amount","pending_amount","days_old"]].copy()
            show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
            show.columns = ["Customer","Date","Category","Amount ₹","Pending ₹","Days Old"]
            st.dataframe(show, use_container_width=True, hide_index=True)
            bulk_flags = st.multiselect(
                "Clear multiple flags",
                dl["id"].tolist(),
                format_func=lambda x: f"#{x} — {dl[dl['id']==x]['customer_name'].iloc[0]}",
                key="reminders_bulk_flags",
            )
            if bulk_flags and st.button("Clear Selected Flags", key="reminders_clear_bulk_flags"):
                cleared = 0
                for rid in bulk_flags:
                    ok, _ = update_record_with_audit(rid, set_fields={"delay_status": 0}, action="bulk_clear_flag")
                    cleared += int(ok)
                invalidate_cache()
                st.success(f"{cleared} flag(s) cleared.")
                st.rerun()
            sc = st.selectbox("Clear flag for:", dl["id"].tolist(), format_func=lambda x: f"#{x} — {dl[dl['id']==x]['customer_name'].values[0]}")
            if st.button("Clear Flag"):
                update_record_with_audit(sc, set_fields={"delay_status":0}, action="clear_flag")
                invalidate_cache(); st.success("Flag cleared."); st.rerun()

    with t3:
        hv = df[df["total_amount"] >= 10000].sort_values("total_amount", ascending=False).head(20).copy()
        if hv.empty:
            st.info("No high-value sales (₹10,000+) yet.")
        else:
            hv["sale_date"]        = hv["sale_date"].dt.strftime("%d %b %Y")
            hv["payment_received"] = hv["payment_received"].map({0:"Pending",1:"Paid"})
            show = hv[["customer_name","sale_date","product_category","total_amount","profit","payment_received"]].copy()
            show.columns = ["Customer","Date","Category","Amount ₹","Profit ₹","Status"]
            st.dataframe(show, use_container_width=True, hide_index=True)

    with t4:
        soon = df[(df["pending_amount"] > 0) & (df["days_old"] >= 7) & (df["days_old"] <= 30) & (df["delay_status"] == 0)].sort_values("days_old", ascending=False)
        if soon.empty:
            st.info("No follow-ups needed in the 7–30 day window.")
        else:
            st.info(f"{len(soon)} sales with pending payments between 7–30 days old.")
            show = soon[["customer_name","customer_phone","sale_date","product_category","pending_amount","days_old"]].copy()
            show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
            show.columns = ["Customer","Phone","Date","Category","Pending ₹","Days Old"]
            st.dataframe(show, use_container_width=True, hide_index=True)


def page_inventory():
    page_header("Inventory", "Stock Management")
    inv_col = get_db()["inventory"]
    t1, t2 = st.tabs(["Current Stock","Add / Update Stock"])

    with t1:
        items = get_inventory_items()
        if not items:
            st.markdown("<div class='empty'><div class='empty-glyph'>◆</div><div>No inventory items yet.</div></div>", unsafe_allow_html=True)
        else:
            inv_df = pd.DataFrame(items)
            q_col = inv_df["quantity"] if "quantity" in inv_df.columns else pd.Series([0] * len(inv_df), index=inv_df.index)
            cost_col = inv_df["cost_price"] if "cost_price" in inv_df.columns else pd.Series([0] * len(inv_df), index=inv_df.index)
            min_col = inv_df["min_stock"] if "min_stock" in inv_df.columns else pd.Series([5] * len(inv_df), index=inv_df.index)
            total_value  = (q_col * cost_col).sum()
            low_stock    = inv_df[q_col <= min_col]
            out_of_stock = inv_df[q_col == 0]
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total SKUs",      len(inv_df))
            m2.metric("Inventory Value", f"₹{total_value:,.0f}")
            m3.metric("Low Stock",       len(low_stock))
            m4.metric("Out of Stock",    len(out_of_stock))
            if not low_stock.empty: st.warning(f"{len(low_stock)} item(s) running low.")
            rule_sm()
            inv_search = st.text_input("Search SKU / Item / Vendor", key="inventory_search")
            cat_f = st.selectbox("Filter by Category", ["All"] + CATEGORIES, key="inventory_category_filter")
            view  = inv_df.copy()
            if cat_f != "All" and "category" in view.columns: view = view[view["category"] == cat_f]
            if inv_search:
                view = view[
                    view.get("sku", pd.Series("", index=view.index)).astype(str).str.contains(inv_search, case=False, na=False)
                    | view.get("name", pd.Series("", index=view.index)).astype(str).str.contains(inv_search, case=False, na=False)
                    | view.get("vendor", pd.Series("", index=view.index)).astype(str).str.contains(inv_search, case=False, na=False)
                ]
            if "quantity" in view.columns and "min_stock" in view.columns:
                view["Status"] = view.apply(lambda r: "Out of Stock" if r["quantity"]==0 else ("Low Stock" if r["quantity"]<=r["min_stock"] else "OK"), axis=1)
            st.dataframe(paginated_slice(view, "inventory_table"), use_container_width=True, hide_index=True)
            labels_df = view[[c for c in ["sku", "name", "category", "sell_price"] if c in view.columns]].copy()
            st.download_button("Download Label CSV", data=labels_df.to_csv(index=False), file_name=f"inventory_labels_{date.today()}.csv", mime="text/csv", use_container_width=True)
            if "category" in inv_df.columns and "quantity" in inv_df.columns:
                cat_stock = inv_df.groupby("category")["quantity"].sum().reset_index()
                fig = px.bar(cat_stock, x="category", y="quantity", title="Stock by Category", color="quantity", color_continuous_scale=[[0,"#C05060"],[0.4,"#2E6FD8"],[1,"#7ADFA0"]])
                styled_fig(fig, 260); st.plotly_chart(fig, use_container_width=True)

    with t2:
        sec("Add or Update Stock Item")
        with st.form("inv_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                item_name = st.text_input("Item Name *", placeholder="e.g. Banarasi Silk Saree")
                item_sku  = st.text_input("SKU / Code",  placeholder="e.g. SAR-001")
                item_cat  = st.selectbox("Category", CATEGORIES)
                item_vend = st.text_input("Vendor")
                variant = st.text_input("Variant", placeholder="Size / color / fabric")
            with c2:
                item_qty  = st.number_input("Quantity *",       min_value=0, step=1)
                item_min  = st.number_input("Min Stock Alert",  min_value=0, step=1, value=5)
                item_cost = st.number_input("Cost Price (₹) *", min_value=0.0, step=50.0, format="%.2f")
                item_mrp  = st.number_input("Selling Price (₹)",min_value=0.0, step=50.0, format="%.2f")
                season_tag = st.text_input("Season / Expiry Tag", placeholder="Summer 2026, Festive, etc.")
            item_notes = st.text_area("Notes", height=55)
            if st.form_submit_button("Save Item", use_container_width=True):
                if not item_name.strip():
                    st.error("Item name is required.")
                else:
                    sku_value = item_sku.strip() or make_sku(item_name.strip(), item_cat)
                    set_fields = {
                        "name": item_name.strip(),
                        "sku": sku_value,
                        "category": item_cat,
                        "vendor": item_vend.strip(),
                        "variant": variant.strip(),
                        "quantity": item_qty,
                        "min_stock": item_min,
                        "cost_price": round(item_cost,2),
                        "sell_price": round(item_mrp,2),
                        "season_tag": season_tag.strip(),
                        "notes": item_notes.strip(),
                        "updated_at": str(datetime.now()),
                    }
                    old = inv_col.find_one({"sku": sku_value}) or {}
                    inv_col.update_one(
                        {"sku": sku_value},
                        {"$set": set_fields, "$setOnInsert": {"created_at": str(datetime.now()), "created_by": current_user()}},
                        upsert=True,
                    )
                    audit_log(sku_value, "inventory_upsert", build_changes(old, set_fields), collection="inventory")
                    st.success(f"'{item_name.strip()}' saved to inventory.")
                    st.rerun()

def page_daily_summary():
    page_header("Daily Summary", "Cash, UPI, Card and Bank Totals")
    df = fetch_all()
    if df.empty:
        st.info("No transactions available.")
        return
    today = date.today()
    c1, c2 = st.columns(2)
    with c1:
        d_from = st.date_input("From", today, key="daily_from")
    with c2:
        d_to = st.date_input("To", today, key="daily_to")
    window = df[(df["sale_date"] >= pd.Timestamp(d_from)) & (df["sale_date"] <= pd.Timestamp(d_to))].copy()
    if window.empty:
        st.info("No records in this period.")
        return

    sales = accounted_sales(window)
    expenses = expense_records(window)
    returns = return_records(window)
    collected = sales["amount_paid"].sum() if not sales.empty else 0
    pending = sales["pending_amount"].sum() if not sales.empty else 0
    expense_total = expenses["selling_price"].sum() if not expenses.empty else 0
    refund_total = returns["amount_paid"].sum() if not returns.empty else 0
    net_cash = collected - expense_total - refund_total

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Collected", f"₹{collected:,.0f}")
    m2.metric("Pending Added", f"₹{pending:,.0f}")
    m3.metric("Expenses", f"₹{expense_total:,.0f}")
    m4.metric("Refunds", f"₹{refund_total:,.0f}")
    m5.metric("Net Inflow", f"₹{net_cash:,.0f}")

    rule_sm()
    sec("Payment Method Breakup")
    payments = window[window["amount_paid"] > 0].groupby("payment_method")["amount_paid"].sum().reset_index()
    if payments.empty:
        st.info("No payments collected in this period.")
    else:
        payments.columns = ["Method", "Amount ₹"]
        st.dataframe(payments, use_container_width=True, hide_index=True)
        fig = px.bar(payments, x="Method", y="Amount ₹", title="Collected by Payment Method", color="Amount ₹", color_continuous_scale=[[0,"#2E6FD8"],[1,"#7ADFA0"]])
        styled_fig(fig, 260)
        st.plotly_chart(fig, use_container_width=True)

    sec("Cash Register Reconciliation")
    expected_cash = float(payments[payments["Method"].eq("Cash")]["Amount ₹"].sum()) if not payments.empty and "Method" in payments.columns else 0.0
    with st.form("cash_reconciliation_form"):
        r1, r2, r3 = st.columns(3)
        with r1:
            actual_cash = st.number_input("Actual Cash Counted (₹)", min_value=0.0, value=expected_cash, step=100.0, format="%.2f")
        with r2:
            variance = round(actual_cash - expected_cash, 2)
            st.metric("Variance", f"₹{variance:,.2f}")
        with r3:
            recon_note = st.text_input("Note")
        if st.form_submit_button("Save Reconciliation", use_container_width=True):
            get_collection("cash_reconciliations").update_one(
                {"date_from": str(d_from), "date_to": str(d_to)},
                {"$set": {"expected_cash": expected_cash, "actual_cash": round(float(actual_cash), 2), "variance": variance, "note": recon_note.strip()[:300], "updated_at": str(datetime.now()), "updated_by": current_user()}},
                upsert=True,
            )
            audit_log(f"{d_from}_{d_to}", "cash_reconciliation", [{"field": "variance", "old_value": None, "new_value": variance}], collection="cash_reconciliations")
            st.success("Cash reconciliation saved.")

    sec("Transactions")
    show = window[["id","transaction_type","customer_name","sale_date","payment_method","amount_paid","pending_amount","notes"]].copy()
    show["sale_date"] = show["sale_date"].dt.strftime("%d %b %Y")
    show.columns = ["ID","Type","Name","Date","Method","Paid ₹","Pending ₹","Notes"]
    st.dataframe(show, use_container_width=True, hide_index=True)

def page_vendor_payables():
    page_header("Vendor Payables", "Supplier Outstanding Tracker")
    df = fetch_all()
    if df.empty:
        st.info("No transaction data available.")
        return
    sales = accounted_sales(df)
    sales = sales[sales["vendor"].astype(str).str.strip() != ""].copy()
    vendor_payments = expense_records(df)
    vendor_payments = vendor_payments[vendor_payments["expense_category"].eq("Vendor Payment") & vendor_payments["vendor"].astype(str).str.strip().ne("")]

    if sales.empty and vendor_payments.empty:
        st.info("Add vendor names to sales or record Vendor Payment expenses to see payables.")
        return

    purchased = sales.assign(cost_total=sales["buying_price"] * sales["quantity"]).groupby("vendor")["cost_total"].sum()
    paid = vendor_payments.groupby("vendor")["selling_price"].sum() if not vendor_payments.empty else pd.Series(dtype=float)
    vendors = sorted(set(purchased.index.tolist()) | set(paid.index.tolist()))
    rows = []
    for vendor in vendors:
        buy_total = float(purchased.get(vendor, 0) or 0)
        paid_total = float(paid.get(vendor, 0) or 0)
        rows.append({"Vendor": vendor, "Purchase Cost ₹": buy_total, "Paid ₹": paid_total, "Outstanding ₹": buy_total - paid_total})
    view = pd.DataFrame(rows).sort_values("Outstanding ₹", ascending=False)

    m1, m2, m3 = st.columns(3)
    m1.metric("Purchase Cost", f"₹{view['Purchase Cost ₹'].sum():,.0f}")
    m2.metric("Paid Vendors", f"₹{view['Paid ₹'].sum():,.0f}")
    m3.metric("Outstanding", f"₹{view['Outstanding ₹'].sum():,.0f}")
    st.dataframe(view.style.format({"Purchase Cost ₹":"₹{:,.0f}","Paid ₹":"₹{:,.0f}","Outstanding ₹":"₹{:,.0f}"}), use_container_width=True, hide_index=True)

    sec("Record Vendor Payment")
    with st.form("vendor_payment_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            vendor = st.selectbox("Vendor", vendors)
        with c2:
            amount = st.number_input("Amount Paid (₹)", min_value=0.0, step=100.0, format="%.2f")
        with c3:
            method = st.selectbox("Method", PAYMENT_METHODS)
        note = st.text_area("Notes", height=60)
        if st.form_submit_button("Save Vendor Payment", use_container_width=True):
            if amount <= 0:
                st.error("Amount must be > 0.")
            else:
                insert_transaction({
                    "id": get_next_id(),
                    "transaction_type": "expense",
                    "customer_name": vendor,
                    "customer_phone": "",
                    "sale_date": str(date.today()),
                    "vendor": vendor,
                    "product_category": "Other",
                    "expense_category": "Vendor Payment",
                    "product_description": "Vendor Payment",
                    "quantity": 1,
                    "buying_price": 0.0,
                    "selling_price": round(float(amount), 2),
                    "amount_paid": round(float(amount), 2),
                    "pending_amount": 0.0,
                    "payment_received": 1,
                    "delay_status": 0,
                    "payment_method": method,
                    "notes": note.strip()[:500],
                    "recorded_by": st.session_state.get("username", "Admin"),
                    "created_at": str(datetime.now()),
                }, action="vendor_payment")
                invalidate_cache()
                st.success("Vendor payment recorded.")
                st.rerun()

def page_staff_audit():
    page_header("Staff & Audit", "Users, Trash, and Operational Control")
    t1, t2, t3, t4, t5 = st.tabs(["Staff Accounts", "Audit Trail", "Trash", "Attendance", "Tasks"])

    with t1:
        if not is_admin_user():
            st.warning("Only admin users can manage staff accounts.")
        else:
            staff_docs = list(get_staff_col().find({}, {"password_hash": 0}).sort("username", 1))
            if staff_docs:
                staff_df = pd.DataFrame(staff_docs)
                if "_id" in staff_df.columns:
                    staff_df["_id"] = staff_df["_id"].astype(str)
                st.dataframe(staff_df, use_container_width=True, hide_index=True)
            else:
                st.info("No staff accounts yet. The configured admin is bootstrapped automatically on login.")

            with st.form("staff_create_form"):
                s1, s2, s3 = st.columns(3)
                with s1:
                    username = st.text_input("Username *")
                    full_name = st.text_input("Full Name")
                with s2:
                    role = st.selectbox("Role", STAFF_ROLES, index=2)
                    active = st.checkbox("Active", value=True)
                with s3:
                    password = st.text_input("Temporary Password *", type="password")
                if st.form_submit_button("Create Staff Account", use_container_width=True):
                    if not username.strip() or not password:
                        st.error("Username and password are required.")
                    elif role not in STAFF_ROLES:
                        st.error("Invalid role.")
                    elif get_staff_col().count_documents({"username_norm": username.strip().lower()}, limit=1):
                        st.error("That username already exists.")
                    else:
                        doc = {
                            "username": username.strip(),
                            "username_norm": username.strip().lower(),
                            "full_name": full_name.strip() or username.strip().title(),
                            "role": role,
                            "password_hash": _hash_password(password),
                            "active": active,
                            "created_at": str(datetime.now()),
                            "created_by": current_user(),
                        }
                        get_staff_col().insert_one(doc)
                        audit_log(username.strip(), "staff_create", [{"field": "staff_user", "old_value": None, "new_value": {"username": username.strip(), "role": role, "active": active}}], collection="staff_users")
                        st.success("Staff account created.")
                        st.rerun()

            users = [d["username"] for d in get_staff_col().find({}, {"username": 1}).sort("username", 1)]
            if users:
                sec("Update Staff")
                target = st.selectbox("Staff User", users, key="staff_update_user")
                target_doc = get_staff_col().find_one({"username": target}) or {}
                u1, u2, u3 = st.columns(3)
                with u1:
                    new_role = st.selectbox("Role", STAFF_ROLES, index=STAFF_ROLES.index(target_doc.get("role", "cashier")) if target_doc.get("role") in STAFF_ROLES else 0, key="staff_update_role")
                with u2:
                    new_active = st.checkbox("Active", value=bool(target_doc.get("active", True)), key="staff_update_active")
                with u3:
                    reset_pw = st.text_input("Reset Password", type="password", key="staff_reset_pw")
                if st.button("Save Staff Changes", key="staff_update_save"):
                    set_fields = {"role": new_role, "active": new_active, "updated_at": str(datetime.now()), "updated_by": current_user()}
                    if reset_pw:
                        set_fields["password_hash"] = _hash_password(reset_pw)
                    old = target_doc.copy()
                    get_staff_col().update_one({"username": target}, {"$set": set_fields})
                    safe_fields = {k: v for k, v in set_fields.items() if k != "password_hash"}
                    audit_log(target, "staff_update", build_changes(old, safe_fields), collection="staff_users")
                    st.success("Staff updated.")
                    st.rerun()

    with t2:
        logs = list(get_audit_col().find({}, {"_id": 0}).sort("timestamp", -1).limit(500))
        if not logs:
            st.info("No audit entries yet.")
        else:
            adf = pd.DataFrame(logs)
            af1, af2, af3 = st.columns(3)
            with af1:
                user_f = st.text_input("User", key="audit_filter_user")
            with af2:
                action_f = st.text_input("Action", key="audit_filter_action")
            with af3:
                record_f = st.text_input("Record ID", key="audit_filter_record")
            if user_f:
                adf = adf[adf["user"].astype(str).str.contains(user_f, case=False, na=False)]
            if action_f:
                adf = adf[adf["action"].astype(str).str.contains(action_f, case=False, na=False)]
            if record_f:
                adf = adf[adf["record_id"].astype(str).str.contains(record_f, case=False, na=False)]
            adf["changes"] = adf["changes"].apply(lambda v: stable_json(v)[:800])
            st.dataframe(paginated_slice(adf, "audit_table", default_size=25), use_container_width=True, hide_index=True)

    with t3:
        trashed = list(get_col().find({"deleted_at": {"$exists": True}}, {"_id": 0}).sort("deleted_at", -1))
        if not trashed:
            st.success("Trash is empty.")
        else:
            trash_df = pd.DataFrame(trashed)
            cols = [c for c in ["id", "customer_name", "sale_date", "product_category", "total_amount", "deleted_at", "deleted_by", "deleted_reason"] if c in trash_df.columns]
            st.dataframe(paginated_slice(trash_df[cols], "trash_table", default_size=10), use_container_width=True, hide_index=True)
            rid = st.selectbox("Recover Transaction", trash_df["id"].tolist(), format_func=lambda x: f"#{x} — {trash_df[trash_df['id']==x]['customer_name'].iloc[0]}")
            if st.button("Recover from Trash", key="recover_trash"):
                ok, status = recover_deleted_transaction(rid)
                if ok:
                    invalidate_cache()
                    st.success("Transaction recovered.")
                    st.rerun()
                elif status == "inventory":
                    st.warning("Could not recover because the matching inventory SKU does not have enough stock to deduct.")
                else:
                    st.warning("Could not recover the transaction.")

    with t4:
        staff_names = [d["username"] for d in get_staff_col().find({"active": True}, {"username": 1}).sort("username", 1)] or [current_user()]
        with st.form("attendance_form"):
            a1, a2, a3, a4 = st.columns(4)
            with a1:
                staff_name = st.selectbox("Staff", staff_names)
            with a2:
                shift_date = st.date_input("Shift Date", date.today())
            with a3:
                shift_start = st.text_input("Start", value="10:00")
            with a4:
                shift_end = st.text_input("End", value="19:00")
            if st.form_submit_button("Save Shift", use_container_width=True):
                get_collection("staff_attendance").insert_one({"staff": staff_name, "shift_date": str(shift_date), "start": shift_start, "end": shift_end, "created_at": str(datetime.now()), "created_by": current_user()})
                audit_log(staff_name, "attendance_save", [{"field": "shift_date", "old_value": None, "new_value": str(shift_date)}], collection="staff_attendance")
                st.success("Shift saved.")
        recent_attendance = list(get_collection("staff_attendance").find({}, {"_id": 0}).sort("created_at", -1).limit(50))
        if recent_attendance:
            st.dataframe(pd.DataFrame(recent_attendance), use_container_width=True, hide_index=True)

    with t5:
        staff_names = [d["username"] for d in get_staff_col().find({"active": True}, {"username": 1}).sort("username", 1)] or [current_user()]
        with st.form("task_form"):
            ta1, ta2, ta3 = st.columns(3)
            with ta1:
                assigned_to = st.selectbox("Assign To", staff_names, key="task_staff")
            with ta2:
                due_date = st.date_input("Due Date", date.today(), key="task_due")
            with ta3:
                status = st.selectbox("Status", ["Open", "In Progress", "Done"], key="task_status")
            task = st.text_area("Task / Follow-up", placeholder="Call customer about pending alteration, collect payment, etc.", height=70)
            if st.form_submit_button("Save Task", use_container_width=True):
                if not task.strip():
                    st.error("Task text is required.")
                else:
                    get_collection("staff_tasks").insert_one({"task": task.strip()[:500], "assigned_to": assigned_to, "due_date": str(due_date), "status": status, "created_at": str(datetime.now()), "created_by": current_user()})
                    audit_log(assigned_to, "task_create", [{"field": "task", "old_value": None, "new_value": task.strip()[:500]}], collection="staff_tasks")
                    st.success("Task saved.")
        tasks = list(get_collection("staff_tasks").find({}, {"_id": 0}).sort("due_date", 1).limit(100))
        if tasks:
            st.dataframe(pd.DataFrame(tasks), use_container_width=True, hide_index=True)

def page_operations_hub():
    page_header("Operations Hub", "Orders, Procurement, CRM, Finance, and Growth")
    t1, t2, t3, t4, t5, t6 = st.tabs(["Billing & Orders", "Procurement", "CRM", "Finance", "Marketing & Reports", "Platform"])

    with t1:
        sec("Quotation / Estimate")
        with st.form("quote_form"):
            q1, q2, q3 = st.columns(3)
            with q1:
                q_customer = st.text_input("Customer *", key="quote_customer")
                q_phone = st.text_input("Phone", key="quote_phone")
            with q2:
                q_valid = st.date_input("Valid Until", date.today() + timedelta(days=15), key="quote_valid")
                q_total = st.number_input("Estimate Total (₹)", min_value=0.0, step=100.0, format="%.2f", key="quote_total")
            with q3:
                q_status = st.selectbox("Status", ["Draft", "Sent", "Accepted", "Converted", "Expired"], key="quote_status")
                q_desc = st.text_input("Items / Notes", key="quote_desc")
            if st.form_submit_button("Save Quotation", use_container_width=True):
                if not q_customer.strip() or q_total <= 0:
                    st.error("Customer and estimate total are required.")
                else:
                    get_collection("quotations").insert_one({"customer": q_customer.strip(), "phone": q_phone.strip(), "valid_until": str(q_valid), "total": round(float(q_total), 2), "status": q_status, "description": q_desc.strip(), "created_at": str(datetime.now()), "created_by": current_user()})
                    audit_log(q_customer.strip(), "quotation_create", [{"field": "total", "old_value": None, "new_value": q_total}], collection="quotations")
                    st.success("Quotation saved.")
        quotes = list(get_collection("quotations").find({"status": {"$ne": "Converted"}}, {"_id": 0}).sort("created_at", -1).limit(50))
        if quotes:
            qdf = pd.DataFrame(quotes)
            st.dataframe(qdf, use_container_width=True, hide_index=True)
            quote_idx = st.selectbox("Convert Quote", list(range(len(quotes))), format_func=lambda i: f"{quotes[i].get('customer')} · ₹{quotes[i].get('total', 0):,.0f}", key="quote_convert")
            if st.button("Convert Selected Quote to Sale", key="quote_convert_btn"):
                q = quotes[int(quote_idx)]
                pending_amt = round(float(q.get("total", 0) or 0), 2)
                insert_transaction({
                    "id": get_next_id(),
                    "transaction_type": "sale",
                    "sale_mode": "quote_conversion",
                    "customer_name": q.get("customer", ""),
                    "customer_phone": q.get("phone", ""),
                    "sale_date": str(date.today()),
                    "vendor": "",
                    "product_category": "Other",
                    "product_description": q.get("description", "Quotation conversion"),
                    "quantity": 1,
                    "buying_price": 0.0,
                    "selling_price": pending_amt,
                    "gross_amount": pending_amt,
                    "amount_paid": 0.0,
                    "pending_amount": pending_amt,
                    "payment_received": 0,
                    "delay_status": 0,
                    "payment_method": "Credit",
                    "notes": "Converted from quotation",
                    "recorded_by": current_user(),
                }, action="quote_convert_sale")
                get_collection("quotations").update_one({"customer": q.get("customer"), "created_at": q.get("created_at")}, {"$set": {"status": "Converted", "converted_at": str(datetime.now()), "converted_by": current_user()}})
                invalidate_cache()
                st.success("Quote converted into a pending sale.")
                st.rerun()

        sec("Tailoring, Alterations, Trial Queue, Layaway, Gift Cards")
        order_col, layaway_col = st.columns(2)
        with order_col:
            with st.form("custom_order_form"):
                o_customer = st.text_input("Customer", key="order_customer")
                order_type = st.selectbox("Order Type", ["Trial Room Queue", "Custom Stitching", "Alteration"], key="order_type")
                measurements = st.text_area("Measurements / Instructions", key="order_measurements")
                tailor = st.text_input("Tailor / Owner", key="order_tailor")
                due = st.date_input("Delivery Date", date.today() + timedelta(days=7), key="order_due")
                status = st.selectbox("Status", ORDER_STATUSES, key="order_status")
                if st.form_submit_button("Save Order", use_container_width=True):
                    get_collection("orders").insert_one({"customer": o_customer.strip(), "order_type": order_type, "measurements": measurements.strip(), "tailor": tailor.strip(), "due_date": str(due), "status": status, "created_at": str(datetime.now()), "created_by": current_user()})
                    audit_log(o_customer.strip(), "order_create", [{"field": "status", "old_value": None, "new_value": status}], collection="orders")
                    st.success("Order saved.")
        with layaway_col:
            with st.form("layaway_gift_form"):
                lg_customer = st.text_input("Customer", key="lg_customer")
                lg_type = st.selectbox("Record Type", ["Layaway / Installment", "Gift Card", "Store Credit"], key="lg_type")
                lg_amount = st.number_input("Value (₹)", min_value=0.0, step=100.0, format="%.2f", key="lg_amount")
                lg_paid = st.number_input("Paid / Issued (₹)", min_value=0.0, step=100.0, format="%.2f", key="lg_paid")
                lg_code = st.text_input("Code / Reference", value=make_sku(lg_customer or "CREDIT", lg_type), key="lg_code")
                if st.form_submit_button("Save Credit Record", use_container_width=True):
                    get_collection("customer_credits").insert_one({"customer": lg_customer.strip(), "type": lg_type, "amount": round(float(lg_amount), 2), "paid": round(float(lg_paid), 2), "balance": max(round(float(lg_amount - lg_paid), 2), 0.0), "code": lg_code.strip(), "created_at": str(datetime.now()), "created_by": current_user()})
                    audit_log(lg_code.strip(), "credit_record_create", [{"field": "balance", "old_value": None, "new_value": max(round(float(lg_amount - lg_paid), 2), 0.0)}], collection="customer_credits")
                    st.success("Credit record saved.")

    with t2:
        sec("Purchase Order / Receiving")
        with st.form("po_form"):
            p1, p2, p3, p4 = st.columns(4)
            with p1:
                po_vendor = st.text_input("Vendor *", key="po_vendor")
                po_sku = st.text_input("SKU", key="po_sku")
            with p2:
                po_item = st.text_input("Item *", key="po_item")
                po_cat = st.selectbox("Category", CATEGORIES, key="po_category")
            with p3:
                po_qty = st.number_input("Quantity", min_value=1, step=1, key="po_qty")
                po_cost = st.number_input("Unit Cost ₹", min_value=0.0, step=50.0, format="%.2f", key="po_cost")
            with p4:
                po_expected = st.date_input("Expected Delivery", date.today() + timedelta(days=7), key="po_expected")
                po_status = st.selectbox("Status", ["Ordered", "Part Received", "Received", "Cancelled"], key="po_status")
            if st.form_submit_button("Save Purchase Order", use_container_width=True):
                sku_value = po_sku.strip() or make_sku(po_item, po_cat)
                get_collection("purchase_orders").insert_one({"vendor": po_vendor.strip(), "sku": sku_value, "item": po_item.strip(), "category": po_cat, "quantity": int(po_qty), "unit_cost": round(float(po_cost), 2), "expected_delivery": str(po_expected), "status": po_status, "created_at": str(datetime.now()), "created_by": current_user()})
                if po_status == "Received":
                    increment_inventory(po_cat, int(po_qty), po_item, po_vendor, po_cost, sku_value)
                audit_log(sku_value, "purchase_order_save", [{"field": "status", "old_value": None, "new_value": po_status}], collection="purchase_orders")
                st.success("Purchase order saved.")

        sec("Transfers, Variants, Return to Vendor")
        pv1, pv2 = st.columns(2)
        with pv1:
            with st.form("stock_transfer_form"):
                tr_sku = st.text_input("SKU", key="transfer_sku")
                tr_qty = st.number_input("Quantity", min_value=1, step=1, key="transfer_qty")
                tr_from = st.text_input("From Location", key="transfer_from")
                tr_to = st.text_input("To Location", key="transfer_to")
                if st.form_submit_button("Log Transfer", use_container_width=True):
                    get_collection("stock_transfers").insert_one({"sku": tr_sku.strip(), "quantity": int(tr_qty), "from": tr_from.strip(), "to": tr_to.strip(), "created_at": str(datetime.now()), "created_by": current_user()})
                    audit_log(tr_sku.strip(), "stock_transfer", [{"field": "quantity", "old_value": None, "new_value": int(tr_qty)}], collection="stock_transfers")
                    st.success("Transfer logged.")
        with pv2:
            with st.form("rtv_form"):
                rtv_sku = st.text_input("SKU", key="rtv_sku")
                rtv_vendor = st.text_input("Vendor", key="rtv_vendor")
                rtv_qty = st.number_input("Qty to Return", min_value=1, step=1, key="rtv_qty")
                rtv_reason = st.text_input("Reason", key="rtv_reason")
                if st.form_submit_button("Return to Vendor", use_container_width=True):
                    matched = decrement_inventory_for_sale("Other", int(rtv_qty), "", rtv_sku.strip())
                    get_collection("returns_to_vendor").insert_one({"sku": rtv_sku.strip(), "vendor": rtv_vendor.strip(), "quantity": int(rtv_qty), "reason": rtv_reason.strip(), "inventory_deducted": matched, "created_at": str(datetime.now()), "created_by": current_user()})
                    audit_log(rtv_sku.strip(), "return_to_vendor", [{"field": "quantity", "old_value": None, "new_value": int(rtv_qty)}], collection="returns_to_vendor")
                    st.success("Return to vendor logged." if matched else "RTV logged, but exact SKU stock was not deducted.")

    with t3:
        sec("Measurement Profile, Preferences, Dates")
        with st.form("crm_profile_form"):
            cp1, cp2, cp3 = st.columns(3)
            with cp1:
                crm_customer = st.text_input("Customer *", key="crm_customer")
                birthday = st.date_input("Birthday", date.today(), key="crm_birthday")
            with cp2:
                anniversary = st.date_input("Anniversary", date.today(), key="crm_anniversary")
                preferences = st.text_input("Style / Color / Category Preferences", key="crm_preferences")
            with cp3:
                wishlist = st.text_input("Wishlist", key="crm_wishlist")
                measurements = st.text_area("Measurements", key="crm_measurements")
            if st.form_submit_button("Save CRM Profile", use_container_width=True):
                get_collection("customer_profiles").update_one(
                    {"customer": crm_customer.strip()},
                    {"$set": {"birthday": str(birthday), "anniversary": str(anniversary), "preferences": preferences.strip(), "wishlist": wishlist.strip(), "measurements": measurements.strip(), "updated_at": str(datetime.now()), "updated_by": current_user()}},
                    upsert=True,
                )
                audit_log(crm_customer.strip(), "crm_profile_save", [{"field": "profile", "old_value": None, "new_value": "updated"}], collection="customer_profiles")
                st.success("CRM profile saved.")

        sec("Campaign Queue, Referrals, Feedback")
        with st.form("campaign_form"):
            ch1, ch2, ch3 = st.columns(3)
            with ch1:
                channel = st.selectbox("Channel", CAMPAIGN_CHANNELS, key="campaign_channel")
                audience = st.selectbox("Audience", ["All customers", "Customers with pending", "Platinum/Gold customers"], key="campaign_audience")
            with ch2:
                campaign_name = st.text_input("Campaign", placeholder="Festival offer, new arrivals", key="campaign_name")
                template = st.text_input("Template", key="campaign_template")
            with ch3:
                message = st.text_area("Message", key="campaign_message")
            if st.form_submit_button("Create Campaign Queue", use_container_width=True):
                df = fetch_all()
                customers = pd.DataFrame()
                if not df.empty:
                    customers = df.groupby("customer_name").agg(phone=("customer_phone", "first"), pending=("pending_amount", "sum"), spend=("total_amount", "sum")).reset_index()
                    if audience == "Customers with pending":
                        customers = customers[customers["pending"] > 0]
                    elif audience == "Platinum/Gold customers":
                        customers = customers[customers["spend"].apply(lambda x: loyalty_for_spend(float(x))["tier"] in ["Gold", "Platinum"])]
                recipients = customers.to_dict("records") if not customers.empty else []
                provider_configured = bool(os.getenv(f"{channel.upper()}_API_KEY"))
                get_collection("campaigns").insert_one({"name": campaign_name.strip(), "channel": channel, "audience": audience, "template": template.strip(), "message": message.strip(), "recipient_count": len(recipients), "send_status": "ready" if provider_configured else "queued_no_provider", "created_at": str(datetime.now()), "created_by": current_user()})
                st.success(f"Campaign queued for {len(recipients)} recipient(s)." + ("" if provider_configured else " No provider key is configured, so nothing was sent automatically."))
        cf1, cf2 = st.columns(2)
        with cf1:
            with st.form("referral_form"):
                ref_customer = st.text_input("Customer", key="ref_customer")
                referred_by = st.text_input("Referred By", key="referred_by")
                reward = st.number_input("Reward ₹", min_value=0.0, step=50.0, format="%.2f", key="ref_reward")
                if st.form_submit_button("Save Referral", use_container_width=True):
                    get_collection("referrals").insert_one({"customer": ref_customer.strip(), "referred_by": referred_by.strip(), "reward": round(float(reward), 2), "created_at": str(datetime.now()), "created_by": current_user()})
                    st.success("Referral saved.")
        with cf2:
            with st.form("feedback_form"):
                fb_customer = st.text_input("Customer", key="fb_customer")
                rating = st.slider("Rating", 1, 5, 5, key="fb_rating")
                feedback = st.text_area("Feedback", key="fb_text")
                if st.form_submit_button("Save Feedback", use_container_width=True):
                    get_collection("feedback").insert_one({"customer": fb_customer.strip(), "rating": int(rating), "feedback": feedback.strip(), "created_at": str(datetime.now()), "created_by": current_user()})
                    st.success("Feedback saved.")

    with t4:
        all_df = fetch_all()
        m = metrics(all_df)
        pl = pd.DataFrame([
            {"Line": "Revenue", "Amount ₹": m["revenue"]},
            {"Line": "Cost-adjusted Gross Profit", "Amount ₹": m["profit"]},
            {"Line": "Expenses", "Amount ₹": -m["expenses"]},
            {"Line": "Net Profit", "Amount ₹": m["net_profit"]},
        ])
        st.dataframe(pl.style.format({"Amount ₹": "₹{:,.0f}"}), use_container_width=True, hide_index=True)
        if not all_df.empty:
            tax_rate = st.number_input("GST Rate % for Export", min_value=0.0, max_value=28.0, value=5.0, step=0.5, key="gst_rate")
            sales = accounted_sales(all_df).copy()
            gst = sales[["id", "sale_date", "customer_name", "total_amount"]].copy()
            gst["HSN"] = ""
            gst["Taxable Value ₹"] = (gst["total_amount"] / (1 + tax_rate / 100)).round(2)
            gst["GST ₹"] = (gst["total_amount"] - gst["Taxable Value ₹"]).round(2)
            gst.columns = ["Invoice", "Date", "Customer", "Invoice Value ₹", "HSN", "Taxable Value ₹", "GST ₹"]
            st.download_button("Download GSTR-1 Style CSV", data=gst.to_csv(index=False), file_name=f"gstr1_export_{date.today()}.csv", mime="text/csv", use_container_width=True)
        sec("Bank Reconciliation")
        with st.form("bank_recon_form"):
            br1, br2, br3 = st.columns(3)
            with br1:
                br_date = st.date_input("Settlement Date", date.today(), key="br_date")
                br_method = st.selectbox("Method", ["UPI", "Card", "Bank Transfer"], key="br_method")
            with br2:
                expected = st.number_input("Expected ₹", min_value=0.0, step=100.0, format="%.2f", key="br_expected")
                actual = st.number_input("Actual ₹", min_value=0.0, step=100.0, format="%.2f", key="br_actual")
            with br3:
                reference = st.text_input("Bank Reference", key="br_ref")
                note = st.text_input("Note", key="br_note")
            if st.form_submit_button("Save Bank Reconciliation", use_container_width=True):
                get_collection("bank_reconciliations").insert_one({"settlement_date": str(br_date), "method": br_method, "expected": round(float(expected), 2), "actual": round(float(actual), 2), "variance": round(float(actual - expected), 2), "reference": reference.strip(), "note": note.strip(), "created_at": str(datetime.now()), "created_by": current_user()})
                st.success("Bank reconciliation saved.")

    with t5:
        df = accounted_sales(fetch_all())
        inv_df = pd.DataFrame(get_inventory_items())
        if df.empty:
            st.info("Sales reports will appear after sales are recorded.")
        else:
            best = df.groupby("product_description").agg(units=("quantity", "sum"), revenue=("total_amount", "sum")).sort_values("units", ascending=False).head(20).reset_index()
            st.dataframe(best, use_container_width=True, hide_index=True)
            clv = df.groupby("customer_name").agg(lifetime_value=("total_amount", "sum"), visits=("id", "count")).reset_index()
            repeat_rate = (clv["visits"].gt(1).mean() * 100) if not clv.empty else 0
            st.metric("Repeat Purchase Rate", f"{repeat_rate:.1f}%")
            staff_perf = df.groupby("recorded_by").agg(sales=("id", "count"), revenue=("total_amount", "sum"), profit=("profit", "sum")).reset_index()
            st.dataframe(staff_perf, use_container_width=True, hide_index=True)
        if not inv_df.empty:
            st.download_button("Export Social Catalog CSV", data=inv_df[[c for c in ["sku", "name", "category", "sell_price", "quantity"] if c in inv_df.columns]].to_csv(index=False), file_name=f"social_catalog_{date.today()}.csv", mime="text/csv", use_container_width=True)
            if "quantity" in inv_df.columns:
                slow = inv_df.sort_values("quantity", ascending=False).tail(20)
                st.dataframe(slow, use_container_width=True, hide_index=True)

    with t6:
        settings = get_collection("settings")
        backup = settings.find_one({"_id": "backup_schedule"}) or {}
        with st.form("platform_settings_form"):
            ps1, ps2, ps3 = st.columns(3)
            with ps1:
                backup_frequency = st.selectbox("Backup Frequency", ["Manual only", "Daily", "Weekly"], index=["Manual only", "Daily", "Weekly"].index(backup.get("frequency", "Manual only")) if backup.get("frequency") in ["Manual only", "Daily", "Weekly"] else 0)
            with ps2:
                accounting_tool = st.selectbox("Accounting Export", ["None", "Tally", "Zoho Books"], index=0)
            with ps3:
                require_2fa = st.checkbox("Require Login 2FA Code", value=bool((settings.find_one({"_id": "security"}) or {}).get("require_2fa", False)))
            offline_note = st.text_input("Offline/PWA Note", value="Use exported checkpoints during outages; full offline sync requires a separate PWA client.")
            if st.form_submit_button("Save Platform Settings", use_container_width=True):
                settings.update_one({"_id": "backup_schedule"}, {"$set": {"frequency": backup_frequency, "updated_at": str(datetime.now()), "updated_by": current_user()}}, upsert=True)
                settings.update_one({"_id": "security"}, {"$set": {"require_2fa": bool(require_2fa), "updated_at": str(datetime.now()), "updated_by": current_user()}}, upsert=True)
                settings.update_one({"_id": "integrations"}, {"$set": {"accounting_tool": accounting_tool, "offline_note": offline_note, "updated_at": str(datetime.now()), "updated_by": current_user()}}, upsert=True)
                audit_log("platform", "platform_settings_update", [{"field": "backup_frequency", "old_value": backup.get("frequency"), "new_value": backup_frequency}], collection="settings")
                st.success("Platform settings saved.")
        st.caption("Scheduled backups, SMS/email sending, PWA offline sync, and Tally/Zoho posting require an external scheduler/provider or client app. This page stores the configuration and avoids pretending an integration ran when no provider is configured.")

# =====================================================
# BACKUP & RESTORE PAGE
# =====================================================

def page_backup_restore():
    # ── Header ────────────────────────────────────────────────────────────
    st.markdown("""
    <div class='bk-header'>
        <span class='bk-header-icon'>🗄️</span>
        <div class='bk-header-title'>Backup &amp; Restore</div>
        <div class='bk-header-sub'>Database Checkpoint Management · Vinay Boutique</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Last backup timestamp from session state ──────────────────────────
    last_backup_ts = st.session_state.get("last_backup_ts", None)
    last_restore_ts = st.session_state.get("last_restore_ts", None)

    # ══════════════════════════════════════════════════════════
    # ROW 1 — Backup & Restore cards side by side
    # ══════════════════════════════════════════════════════════
    col_bk, col_re = st.columns(2, gap="large")

    # ── Backup card ───────────────────────────────────────────────────────
    with col_bk:
        st.markdown("""
        <div class='bk-card' style='animation-delay:0s'>
            <div class='bk-card-icon bk-icon-blue'>💾</div>
            <div class='bk-card-title'>Backup Database</div>
            <div class='bk-card-desc'>
                Export a complete checkpoint of all sales, customers, and inventory data.
                Download as Excel for safekeeping or migration.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        df = fetch_all()

        # CSV checkpoint
        csv_data = df.assign(sale_date=df["sale_date"].astype(str)).to_csv(index=False) if not df.empty else "No data"
        st.download_button(
            label="⬇️  Download CSV Checkpoint",
            data=csv_data,
            file_name=f"boutique_checkpoint_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
        st.caption("Includes all sales, customer, and payment records")

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

        # Excel checkpoint
        if not df.empty:
            excel_data = to_excel(df)
            ts_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            st.download_button(
                label="⬇️  Download Excel Checkpoint",
                data=excel_data,
                file_name=f"boutique_checkpoint_{ts_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("Formatted spreadsheet with column headers and styling")
        else:
            st.info("No data to export yet.")

        if last_backup_ts:
            st.markdown(f"<div class='bk-ts'>Last export: {last_backup_ts}</div>", unsafe_allow_html=True)

        if st.button("📋  Record Manual Backup Note", use_container_width=True):
            ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
            st.session_state.last_backup_ts = ts
            st.success(f"✓ Manual backup noted at {ts}")
            st.rerun()

    # ── Restore card ──────────────────────────────────────────────────────
    with col_re:
        st.markdown("""
        <div class='bk-card' style='animation-delay:0.1s'>
            <div class='bk-card-icon bk-icon-green'>♻️</div>
            <div class='bk-card-title'>Restore from Checkpoint</div>
            <div class='bk-card-desc'>
                Upload a previously exported CSV checkpoint to restore records.
                Existing data will be preserved — only new records are added.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Upload Checkpoint File",
            type=["csv", "xlsx", "xls"],
            help="Upload a CSV or Excel file previously exported from this application",
            label_visibility="visible",
        )
        st.caption("200 MB max · CSV or Excel format")

        if last_restore_ts:
            st.markdown(
                f"<span class='bk-status-badge bk-status-ok'>✓ Last restored: {last_restore_ts}</span>",
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)

        if uploaded is not None:
            try:
                file_name = uploaded.name.lower()
                restore_df = pd.read_excel(uploaded) if file_name.endswith((".xlsx", ".xls")) else pd.read_csv(uploaded)
                row_count = len(restore_df)
                col_count = len(restore_df.columns)

                st.markdown(f"""
                <div class='bk-card' style='animation-delay:0.15s; border-color: rgba(37,99,235,0.3); margin-top:0.8rem'>
                    <div class='bk-card-title' style='font-size:0.85rem'>📊 File Preview</div>
                    <div class='bk-card-desc'>
                        <b>{row_count}</b> records · <b>{col_count}</b> columns detected<br>
                        Columns: {', '.join(restore_df.columns[:6].tolist())}{' …' if col_count > 6 else ''}
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.dataframe(restore_df.head(5), use_container_width=True, hide_index=True)

                st.warning(
                    "⚠️  This will **insert** records from the checkpoint into the live database. "
                    "Duplicate entries may result if records already exist.",
                )

                confirm = st.checkbox("I understand — proceed with restore")
                if confirm:
                    if st.button("🔄  Restore Database from Checkpoint", use_container_width=True):
                        progress_placeholder = st.empty()
                        progress_placeholder.markdown(
                            "<div style='background:var(--bg-2);border-radius:999px;overflow:hidden;height:6px;margin:0.5rem 0'>"
                            "<div class='bk-progress-bar'></div></div>",
                            unsafe_allow_html=True,
                        )
                        status_msg = st.empty()
                        status_msg.info("Restoring records…")

                        # ── Actual restore logic ──────────────────────────
                        inserted = 0
                        skipped  = 0
                        errors   = []

                        # ── Column name normalisation ─────────────────────
                        # The app exports title-cased columns (e.g. "Customer Name")
                        # but internally uses snake_case (e.g. "customer_name").
                        # Build a mapping: title-cased export name → internal name.
                        EXPORT_TO_INTERNAL = {
                            "Id":                  "id",
                            "Transaction Type":    "transaction_type",
                            "Customer Name":       "customer_name",
                            "Customer Phone":      "customer_phone",
                            "Sale Date":           "sale_date",
                            "Product Category":    "product_category",
                            "Expense Category":    "expense_category",
                            "Product Description": "product_description",
                            "Vendor":              "vendor",
                            "Quantity":            "quantity",
                            "Buying Price":        "buying_price",
                            "Selling Price":       "selling_price",
                            "Total Amount":        "total_amount",
                            "Profit":              "profit",
                            "Profit Margin":       "margin",
                            "Profit Margin %":     "margin",
                            "Amount Paid":         "amount_paid",
                            "Pending Amount":      "pending_amount",
                            "Payment Status":      "_payment_status_str",  # converted below
                            "Status":              "_payment_status_str",
                            "Delayed":             "_delayed_str",          # converted below
                            "Payment Method":      "payment_method",
                            "Notes":               "notes",
                            "Created At":          "created_at",
                        }
                        # Rename columns that match the export format; leave unknown ones as-is
                        restore_df = restore_df.rename(
                            columns={k: v for k, v in EXPORT_TO_INTERNAL.items() if k in restore_df.columns}
                        )
                        # Also lowercase any remaining columns that weren't renamed
                        restore_df.columns = [str(c).lower().replace(" ", "_") for c in restore_df.columns]

                        required_cols = {"customer_name", "selling_price"}
                        if not required_cols.issubset(set(restore_df.columns)):
                            progress_placeholder.empty()
                            status_msg.empty()
                            st.error(
                                f"Invalid checkpoint file. Required columns missing: "
                                f"{required_cols - set(restore_df.columns)}"
                            )
                        else:
                            for _, row in restore_df.iterrows():
                                try:
                                    doc = row.dropna().to_dict()

                                    # Convert "Payment Status" string → payment_received int
                                    if "_payment_status_str" in doc:
                                        ps = str(doc.pop("_payment_status_str")).strip().lower()
                                        doc["payment_received"] = 1 if ps in ("paid", "received", "1") else 0

                                    # Convert "Delayed" string → delay_status int
                                    if "_delayed_str" in doc:
                                        dl = str(doc.pop("_delayed_str")).strip().lower()
                                        doc["delay_status"] = 1 if dl in ("yes", "true", "1") else 0

                                    # Derive payment_received from pending_amount if not set
                                    if "payment_received" not in doc:
                                        pending = float(doc.get("pending_amount", 0) or 0)
                                        doc["payment_received"] = 0 if pending > 0 else 1

                                    # Default delay_status
                                    if "delay_status" not in doc:
                                        doc["delay_status"] = 0

                                    # Normalise numeric types
                                    for num_col in ["buying_price", "selling_price", "total_amount", "amount_paid", "pending_amount", "quantity", "profit", "margin"]:
                                        if num_col in doc:
                                            try:
                                                doc[num_col] = float(doc[num_col])
                                            except (ValueError, TypeError):
                                                doc.pop(num_col, None)
                                    for int_col in ["payment_received", "delay_status"]:
                                        if int_col in doc:
                                            doc[int_col] = int(doc[int_col])
                                    if "sale_date" in doc:
                                        try:
                                            doc["sale_date"] = str(pd.to_datetime(doc["sale_date"]).date())
                                        except Exception:
                                            doc["sale_date"] = str(doc["sale_date"])

                                    doc["transaction_type"] = str(doc.get("transaction_type", "sale") or "sale").lower().replace(" ", "_")
                                    dedupe_key = record_fingerprint(doc)
                                    if get_col().count_documents(active_filter({"dedupe_key": dedupe_key}, include_deleted=True), limit=1) > 0:
                                        skipped += 1
                                        continue
                                    doc["dedupe_key"] = dedupe_key

                                    # Drop the old exported id — assign a fresh one
                                    source_id = doc.pop("id", None)
                                    if source_id is not None:
                                        doc["restore_source_id"] = source_id
                                    doc["id"]          = get_next_id()
                                    doc["restored_at"] = str(datetime.now())
                                    insert_transaction(doc, action="restore_insert")
                                    inserted += 1
                                except Exception as e:
                                    skipped += 1
                                    errors.append(str(e))

                            invalidate_cache()
                            progress_placeholder.empty()
                            status_msg.empty()
                            ts = datetime.now().strftime("%d %b %Y, %I:%M %p")
                            st.session_state.last_restore_ts = ts

                            if inserted > 0:
                                st.success(
                                    f"✅  Restore complete — **{inserted}** records inserted"
                                    + (f", {skipped} skipped." if skipped else ".")
                                )
                            if errors:
                                with st.expander(f"⚠️  {len(errors)} row(s) had errors"):
                                    for err in errors[:10]:
                                        st.caption(err)
                            st.rerun()

            except Exception as e:
                st.error(f"Could not read file: {e}")

    # ══════════════════════════════════════════════════════════
    # ROW 2 — Database stats card
    # ══════════════════════════════════════════════════════════
    rule()
    st.markdown("""
    <div class='bk-card' style='animation-delay:0.2s'>
        <div class='bk-card-title'>📈  Current Database Status</div>
        <div class='bk-card-desc'>Live snapshot of records in the database</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    df2 = fetch_all()
    m = metrics(df2)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Records",  m["sales"])
    c2.metric("Total Revenue",  f"₹{m['revenue']:,.0f}")
    c3.metric("Unique Customers", m["customers"])
    c4.metric("Pending Payments", f"₹{m['pending']:,.0f}")
    c5.metric("Data Health",    "✓ Live" if m["sales"] > 0 else "Empty")

    rule_sm()
    st.caption(f"Database last queried: {datetime.now().strftime('%d %b %Y, %I:%M:%S %p')}  ·  Boutique Manager v2.0")


# =====================================================
# MAIN
# =====================================================

def main():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "theme" not in st.session_state:
        st.session_state.theme = "system"

    # Apply light mode CSS overrides if needed
    inject_theme()

    if not enforce_session_timeout():
        page_add_sale(public=True)
        render_admin_login_strip()
        return

    if not is_admin():
        page_add_sale(public=True)
        render_admin_login_strip()
        return

    page = sidebar()

    if   "Dashboard"   in page: page_dashboard()
    elif "Add Sale"    in page: page_add_sale(public=False)
    elif "Review"      in page: page_review()
    elif "Update"      in page: page_update()
    elif "Customer"    in page: page_customers()
    elif "Analytics"   in page: page_analytics()
    elif "Reminders"   in page: page_reminders()
    elif "Daily Cash"  in page: page_daily_summary()
    elif "Vendor"      in page: page_vendor_payables()
    elif "Inventory"   in page: page_inventory()
    elif "Operations"  in page: page_operations_hub()
    elif "Staff"       in page: page_staff_audit()
    elif "Backup"      in page: page_backup_restore()
    elif "Logout"      in page:
        logout_user()
        st.rerun()

if __name__ == "__main__":
    main()
